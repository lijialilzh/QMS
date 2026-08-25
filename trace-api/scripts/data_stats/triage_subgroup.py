# 亚组统计脚本（离线）。详见 docs/function_docs/100_数据文件管理.md。
# 读取病例分布 Excel（可含 gt/pred/dice），按现有列分组，缺列则跳过。全产品通用。

import argparse
import os
import sys

import pandas as pd


def read_table(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return pd.DataFrame(list(rows[1:]), columns=header)


def _col(df, *names):
    mapping = {str(c).strip().lower(): c for c in df.columns}
    for name in names:
        hit = mapping.get(str(name).strip().lower())
        if hit is not None:
            return hit
    return None


def cal_95CI(tp, pos, tn, neg):
    def _ci(success, n):
        if not n:
            return "none"
        try:
            import statsmodels.api as sm
            lo, hi = sm.stats.proportion_confint(success, n, alpha=0.05, method="normal")
            return str(round(success / n, 3)) + str(tuple(round(i, 3) for i in (lo, hi)))
        except Exception:
            p = success / n
            se = (p * (1 - p) / n) ** 0.5
            return str(round(p, 3)) + str((round(max(0, p - 1.96 * se), 3), round(min(1, p + 1.96 * se), 3)))

    return _ci(tp, pos), _ci(tn, neg)


def get_cases(tmp_df):
    result = []
    for i in tmp_df.groups:
        result.append({"interval": str(i), "cases": tmp_df.groups[i].size})
    return result


def main():
    parser = argparse.ArgumentParser(description="按亚组统计阳性/阴性、灵敏度特异度、Dice（全产品通用）")
    parser.add_argument("--input", required=True, help="病例分布 xlsx（data_statistics_multi 输出，可含 gt/pred/dice）")
    parser.add_argument("--output", default="", help="输出 xlsx，默认 <input>_result.xlsx")
    parser.add_argument("--group-cols", default="SEX,DEVICE,state,KVP,THICKNESS,ConvolutionKernel",
                        help="离散分组列，逗号分隔；表中不存在的列自动跳过")
    parser.add_argument("--age-bins", default="0,18,40,60,110", help="AGE 分段边界，逗号分隔")
    args = parser.parse_args()
    if not os.path.isfile(args.input):
        sys.exit("输入文件不存在: %s" % args.input)
    df = read_table(args.input)
    id_col = _col(df, "TXID", "PatientID") or df.columns[0]
    gt_col = _col(df, "gt")
    pred_col = _col(df, "pred")
    dice_col = _col(df, "dice")
    has_label = gt_col is not None and pred_col is not None
    group_cols = []
    for raw in args.group_cols.split(","):
        hit = _col(df, raw.strip())
        if hit is not None:
            group_cols.append(hit)
    age_col = _col(df, "AGE", "年龄")
    age_bins = [float(x) for x in args.age_bins.split(",") if str(x).strip()]
    triage_result = []
    dice_result = []

    def count_id(mask):
        return df.loc[mask, id_col].count()

    if has_label:
        for key in group_cols:
            for factor in df.groupby(key).groups.keys():
                pos = count_id((df[key] == factor) & (df[gt_col] == 1))
                neg = count_id((df[key] == factor) & (df[gt_col] == 0))
                tp = count_id((df[key] == factor) & (df[gt_col] == 1) & (df[pred_col] == 1))
                tn = count_id((df[key] == factor) & (df[gt_col] == 0) & (df[pred_col] == 0))
                sen, spe = cal_95CI(tp, pos, tn, neg)
                triage_result.append({"Item": key, "Catgory": factor, "pos_cases": pos, "neg_cases": neg, "Sen": sen, "Spe": spe})
                if dice_col is not None:
                    dice_df = df[(df[gt_col] == 1) & (df[pred_col] == 1) & (df[key] == factor)][dice_col]
                    dice_result.append({"Item": key, "Catgory": factor, "cases": tp, "Mean": dice_df.mean(), "Std": dice_df.std()})
        if age_col is not None and len(age_bins) >= 2:
            df_pos = df[df[gt_col] == 1]
            df_tp = df[(df[gt_col] == 1) & (df[pred_col] == 1)]
            df_neg = df[df[gt_col] == 0]
            df_tn = df[(df[gt_col] == 0) & (df[pred_col] == 0)]
            result_pos = get_cases(df_pos.groupby(pd.cut(df_pos[age_col], age_bins)))
            result_neg = get_cases(df_neg.groupby(pd.cut(df_neg[age_col], age_bins)))
            result_tp = get_cases(df_tp.groupby(pd.cut(df_tp[age_col], age_bins)))
            result_tn = get_cases(df_tn.groupby(pd.cut(df_tn[age_col], age_bins)))
            for i in range(len(result_pos)):
                pos, tp, tn, neg = result_pos[i]["cases"], result_tp[i]["cases"], result_tn[i]["cases"], result_neg[i]["cases"]
                sen, spe = cal_95CI(tp, pos, tn, neg)
                triage_result.append({"Item": age_col, "Catgory": result_pos[i]["interval"], "pos_cases": pos, "neg_cases": neg, "Sen": sen, "Spe": spe})
            if dice_col is not None:
                tmp_dice = df_tp.groupby(pd.cut(df_tp[age_col], age_bins))[dice_col]
                mean_s, std_s = tmp_dice.mean(), tmp_dice.std()
                for i, cat in enumerate(mean_s.index):
                    dice_result.append({"Item": age_col, "Catgory": str(cat), "cases": result_tp[i]["cases"], "Mean": mean_s.iloc[i], "Std": std_s.iloc[i]})
        pos = count_id(df[gt_col] == 1)
        neg = count_id(df[gt_col] == 0)
        tp = count_id((df[gt_col] == 1) & (df[pred_col] == 1))
        tn = count_id((df[gt_col] == 0) & (df[pred_col] == 0))
        sen, spe = cal_95CI(tp, pos, tn, neg)
        triage_result.append({"Item": "All", "Catgory": "", "pos_cases": pos, "neg_cases": neg, "Sen": sen, "Spe": spe})
        if dice_col is not None:
            dice_df = df[(df[gt_col] == 1) & (df[pred_col] == 1)][dice_col]
            dice_result.append({"Item": "All", "Catgory": "", "cases": tp, "Mean": dice_df.mean(), "Std": dice_df.std()})
    else:
        for key in group_cols:
            for factor, grp in df.groupby(key):
                triage_result.append({"Item": key, "Catgory": factor, "pos_cases": len(grp), "neg_cases": 0, "Sen": "", "Spe": ""})

    out = args.output or os.path.splitext(args.input)[0] + "_result.xlsx"
    from openpyxl import Workbook
    wb = Workbook()

    def write_df(ws, df, cols):
        ws.append(cols)
        for _, row in df[cols].iterrows():
            ws.append([row[c] if pd.notna(row[c]) else "" for c in cols])

    result_df = pd.DataFrame(triage_result)
    dice_df = pd.DataFrame(dice_result)
    first = True
    if not result_df.empty:
        ws = wb.active
        ws.title = "统计结果"
        write_df(ws, result_df, [c for c in ["Item", "Catgory", "pos_cases", "neg_cases", "Sen", "Spe"] if c in result_df.columns])
        first = False
    if not dice_df.empty:
        ws = wb.active if first else wb.create_sheet("Dice结果")
        if first:
            ws.title = "Dice结果"
        write_df(ws, dice_df, ["Item", "Catgory", "cases", "Mean", "Std"])
        first = False
    if not result_df.empty:
        ws = wb.create_sheet("设备分布")
        write_df(ws, result_df, [c for c in ["Item", "Catgory", "pos_cases", "neg_cases"] if c in result_df.columns])
    if first:
        wb.active.title = "统计结果"
    wb.save(out)
    print("wrote", out)


if __name__ == "__main__":
    main()
