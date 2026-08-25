# 把病例分布 Excel 汇总成「因素 / 类别 / 序列数 / 占比」，供数据库统计表展示。
# 全产品通用：只统计表中实际存在的列。

import argparse
import os
import sys
from datetime import date

import pandas as pd


def read_table(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return pd.DataFrame(list(rows[1:]), columns=header)


def _col(df, *names):
    mapping = {str(c).strip().lower(): c for c in df.columns}
    for name in names:
        hit = mapping.get(str(name).strip().lower())
        if hit is not None:
            return hit
    return None


def add_factor(rows, item, category, count, total):
    ratio = (count / total) if total else 0
    rows.append([item, category, int(count), round(float(ratio), 6)])


def main():
    parser = argparse.ArgumentParser(description="汇总病例分布为数据库统计表格式 Excel")
    parser.add_argument("--input", required=True, help="病例分布 xlsx")
    parser.add_argument("--output", default="", help="输出 xlsx")
    parser.add_argument("--title", default="数据库统计表", help="表头标题")
    parser.add_argument("--data-type", default="", help="数据类型，如 胸部CTPA")
    parser.add_argument("--disease", default="", help="疾病构成")
    parser.add_argument("--age-bins", default="0,18,40,60,110", help="年龄分段")
    args = parser.parse_args()
    if not os.path.isfile(args.input):
        sys.exit("输入文件不存在: %s" % args.input)
    df = read_table(args.input)
    total = len(df)
    sex_col = _col(df, "SEX", "性别")
    age_col = _col(df, "AGE", "年龄")
    device_col = _col(df, "DEVICE", "Manufacturer", "设备")
    kvp_col = _col(df, "KVP")
    thick_col = _col(df, "THICKNESS", "层厚")
    hosp_col = _col(df, "医院", "hospital", "site")
    dist_rows = [["因素", "类别", "序列数", "占比"]]
    if sex_col is not None:
        first = True
        for cat, grp in df.groupby(sex_col, dropna=False):
            add_factor(dist_rows, "性别" if first else "", cat if pd.notna(cat) else "", len(grp), total)
            first = False
    if age_col is not None:
        bins = [float(x) for x in args.age_bins.split(",") if str(x).strip()]
        if len(bins) >= 2:
            first = True
            grouped = df.groupby(pd.cut(df[age_col], bins), dropna=False)
            for cat, grp in grouped:
                add_factor(dist_rows, "年龄" if first else "", str(cat), len(grp), total)
                first = False
    if device_col is not None:
        first = True
        for cat, grp in df.groupby(device_col, dropna=False):
            add_factor(dist_rows, "设备" if first else "", cat if pd.notna(cat) else "", len(grp), total)
            first = False
    if kvp_col is not None:
        first = True
        for cat, grp in df.groupby(kvp_col, dropna=False):
            add_factor(dist_rows, "KVP" if first else "", cat if pd.notna(cat) else "", len(grp), total)
            first = False
    if thick_col is not None:
        first = True
        for cat, grp in df.groupby(thick_col, dropna=False):
            add_factor(dist_rows, "层厚" if first else "", cat if pd.notna(cat) else "", len(grp), total)
            first = False
    hosp_n = df[hosp_col].nunique() if hosp_col is not None else ""
    header = [
        [args.title, "", "", ""],
        ["", "", "", ""],
        ["统计人", "", "统计日期", date.today().isoformat()],
        ["数据总量（序列）", total, "数据类型", args.data_type],
        ["疾病构成", args.disease, "医院数量", hosp_n],
        ["数据分布", "", "", ""],
    ]
    grid = header + dist_rows
    out = args.output or os.path.splitext(args.input)[0] + "_stats_table.xlsx"
    out_df = pd.DataFrame(grid)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "数据分布"
    for r_i, row in enumerate(grid, 1):
        for c_i, val in enumerate(row, 1):
            ws.cell(r_i, c_i, val)
    wb.save(out)
    print("wrote", out, "total", total)


if __name__ == "__main__":
    main()
