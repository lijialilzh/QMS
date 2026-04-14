rows = [
    {
        "name": "name1",
        "phone": "phone1",
        "email": "email1",
        "address": "address1"
    },
    {
        "name": "name2",
        "phone": "phone2",
        "email": "email2",
        "address": "address2"
    },
    {
        "name": "name3",
        "phone": "phone3",
        "email": "email3",
        "address": "address3",
        "nodes": [
            {
                "cname": "child4",
                "class": "class4",
            },
            {
                "cname": "child5",
                "class": "class5",
            },
            {
                "cname": "child6",
                "class": "class6",
            },
            {
                "cname": "child7",
                "class": "class7",
            },
            {
                "cname": "child8",
                "class": "class8",
            }
        ],
    },
    {
        "name": "name4",
        "phone": "phone4",
        "email": "email4",
        "address": "address4"
    }
]

from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active
ws.title = "主表+内嵌子表"

# 表头
headers = ["name", "phone", "cname", "class", "email", "address"]
ws.append(headers)

# 自动换行 + 顶部对齐
align = Alignment(vertical='center', wrap_text=True)

# 写入 4 行主数据
for data in rows:
    name = data.get("name", "")
    phone = data.get("phone", "")
    email = data.get("email", "")
    address = data.get("address", "")
    nodes = data.get("nodes", [])

    # 拼接 cname 多行
    cname_lines = "\n".join([node.get("cname", "") for node in nodes])
    # 拼接 class 多行
    class_lines = "\n".join([node.get("class", "") for node in nodes])

    # 写入一行（主数据只占一行，内嵌表格用换行实现）
    ws.append([name, phone, cname_lines, class_lines, email, address])

# 设置所有单元格自动换行 + 顶部对齐
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = align

# 保存
wb.save("内嵌子表格.xlsx")
print("内嵌子表格.xlsx")