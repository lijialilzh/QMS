import logging
import io
import os
import re
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd
from ..model.prod_haz import ProdHaz
from ..obj.vobj_haz import HazObj
from ..model.haz import Haz
from ..model.rcm import Rcm
from ..obj.tobj_haz import HazForm
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


export_columns = [
    "code",
    "source",
    "event",
    "situation",
    "damage",

    "init_rate",
    "init_degree",
    "init_level",

    "deal",
    "rcms",
    "evidence",

    "cur_rate",
    "cur_degree",
    "cur_level",

    "benefit_flag",
    "category"
]

class Server(object):
    @staticmethod
    def __to_str(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def __to_int(value):
        if value is None or value == "":
            return None
        try:
            return int(float(value))
        except Exception:
            return None

    @staticmethod
    def __to_flag(value):
        if value is None:
            return 0
        text = str(value).strip().lower()
        if text in ("y", "yes", "是", "1", "true"):
            return 1
        return 0

    @staticmethod
    def __extract_rcm_codes(rcms_value, deal_value):
        codes = []
        seen = set()

        def add(code):
            key = code.upper()
            if key and key not in seen:
                seen.add(key)
                codes.append(key)

        rcms_text = "" if rcms_value is None else str(rcms_value).strip()
        if rcms_text:
            for item in re.split(r"[,，;；\s]+", rcms_text):
                item = item.strip()
                if item:
                    add(item)
        else:
            deal_text = "" if deal_value is None else str(deal_value)
            for code in re.findall(r"RCM\d+", deal_text, flags=re.IGNORECASE):
                add(code)
        return codes

    @staticmethod
    def __load_rcm_map():
        return {
            (r.code or "").upper(): (r.description or "")
            for r in db.session.execute(select(Rcm)).scalars().all()
        }

    def __resolve_deal_by_rcms(self, rcms_value, rcm_map):
        codes = self.__extract_rcm_codes(rcms_value, None)
        if not codes:
            return None
        lines = []
        for code in codes:
            desc = (rcm_map.get(code) or "").strip()
            lines.append(desc if desc else code)
        return "\n".join(lines)

    @staticmethod
    def __read_cell_value(cell, ws_raw):
        value = cell.value
        if value is not None and str(value).strip() != "":
            return value
        raw = ws_raw.cell(row=cell.row, column=cell.column).value
        if isinstance(raw, str) and raw.startswith("="):
            return raw
        return value

    async def add_haz(self, form: HazForm):
        try:
            sql = select(func.count(Haz.id)).where(Haz.code == form.code)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            
            row = Haz(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def import_hazs(self, file):
        try:
            bys = await file.read()
            wb = load_workbook(io.BytesIO(bys), data_only=True)
            wb_raw = load_workbook(io.BytesIO(bys), data_only=False)
            ws = wb[wb.sheetnames[0]]
            ws_raw = wb_raw[wb_raw.sheetnames[0]]
            rcm_map = {
                (r.code or "").upper(): (r.description or "")
                for r in db.session.execute(select(Rcm)).scalars().all()
            }
            affected = 0
            for row in ws.iter_rows(min_row=3):
                vals = [self.__read_cell_value(cell, ws_raw) for cell in row[:len(export_columns)]]
                if not vals or all((v is None or str(v).strip() == "") for v in vals):
                    continue
                code = self.__to_str(vals[0])
                if not code:
                    continue
                data = dict(
                    code=code,
                    source=self.__to_str(vals[1]),
                    event=self.__to_str(vals[2]),
                    situation=self.__to_str(vals[3]),
                    damage=self.__to_str(vals[4]),
                    init_rate=self.__to_int(vals[5]),
                    init_degree=self.__to_str(vals[6]),
                    init_level=self.__to_str(vals[7]),
                    deal=self.__to_str(vals[8]),
                    rcms=self.__to_str(vals[9]),
                    evidence=self.__to_str(vals[10]),
                    cur_rate=self.__to_int(vals[11]),
                    cur_degree=self.__to_str(vals[12]),
                    cur_level=self.__to_str(vals[13]),
                    benefit_flag=self.__to_flag(vals[14]),
                    category=self.__to_str(vals[15]),
                )
                rcm_codes = self.__extract_rcm_codes(vals[9], vals[8])
                if rcm_codes:
                    deal_lines = []
                    for rcm_code in rcm_codes:
                        desc = (rcm_map.get(rcm_code) or "").strip()
                        deal_lines.append(desc if desc else rcm_code)
                    data["deal"] = "\n".join(deal_lines)
                    data["rcms"] = ",".join(rcm_codes)
                existed = db.session.execute(select(Haz).where(Haz.code == code)).scalars().first()
                if existed:
                    for k, v in data.items():
                        setattr(existed, k, v)
                else:
                    db.session.add(Haz(**data))
                affected += 1
            db.session.commit()
            return Resp.resp_ok(data={"count": affected})
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_haz(self, id):
        db.session.execute(delete(Haz).where(Haz.id == id))
        db.session.execute(delete(ProdHaz).where(ProdHaz.haz_id == id))
        db.session.commit()
        return Resp.resp_ok()
   
    async def update_haz(self, form: HazForm):
        try:
            sql = select(Haz).where(Haz.id == form.id)
            row:Haz = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def get_haz(self, id:str):
        sql = select(Haz).where(Haz.id == id)
        row = db.session.execute(sql).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        obj = HazObj(**row.dict())
        resolved = self.__resolve_deal_by_rcms(row.rcms, self.__load_rcm_map())
        if resolved is not None:
            obj.deal = resolved
        return Resp.resp_ok(data=obj)

    async def list_haz(self, export = False, fuzzy: str = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 
    
        sql = select(Haz)
        if fuzzy:
            sql = sql.where(
                or_(
                    Haz.code.like(f"%{fuzzy}%"),
                    Haz.source.like(f"%{fuzzy}%"),
                    Haz.event.like(f"%{fuzzy}%"),
                    Haz.situation.like(f"%{fuzzy}%"),
                    Haz.damage.like(f"%{fuzzy}%"),

                    Haz.init_degree.like(f"%{fuzzy}%"),
                    Haz.init_level.like(f"%{fuzzy}%"),

                    Haz.deal.like(f"%{fuzzy}%"),
                    Haz.rcms.like(f"%{fuzzy}%"),
                    Haz.evidence.like(f"%{fuzzy}%"),

                    Haz.cur_degree.like(f"%{fuzzy}%"),
                    Haz.cur_level.like(f"%{fuzzy}%"),

                    Haz.category.like(f"%{fuzzy}%")
                )
            )

        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.offset(page_size * page_index).limit(page_size).order_by(Haz.code)
        rows: list[Haz] = db.session.execute(sql).scalars().all()
        rcm_map = self.__load_rcm_map()
        objs = []
        for row in rows:
            obj = HazObj(**row.dict())
            resolved = self.__resolve_deal_by_rcms(row.rcms, rcm_map)
            if resolved is not None:
                obj.deal = resolved
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))
    
    async def export_hazs(self, output, *args, **kwargs):
        resp = await self.list_haz(export=True, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_haz.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 3):
            obj = row.dict()
            for cidx, key in enumerate(export_columns, 1):
                value = obj.get(key)
                if key == "benefit_flag":
                    value = ts("e_yes") if value else ts("e_no")
                ws.cell(row=ridx, column=cidx, value=value)
                ws.cell(row=ridx, column=cidx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        wb.save(output)
        output.seek(0)
        