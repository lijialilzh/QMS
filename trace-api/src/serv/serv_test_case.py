import logging
import os
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from openpyxl import load_workbook

from ..model.product import Product
from ..model.test_set import TestSet
from ..obj.vobj_test_case import TestCaseObj
from ..model.test_case import TestCase
from ..obj.tobj_test_case import TestCaseForm
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):

    async def add_test_case(self, form: TestCaseForm):
        try:
            sql = select(func.count(TestCase.id)).where(TestCase.code == form.code)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            
            row = TestCase(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_test_case(self, id):
        db.session.execute(delete(TestCase).where(TestCase.id == id))
        db.session.commit()
        return Resp.resp_ok()
   
    async def update_test_case(self, form: TestCaseForm):
        try:
            sql = select(TestCase).where(TestCase.id == form.id)
            row:TestCase = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def get_test_case(self, id:str):
        sql = select(TestCase).where(TestCase.id == id)
        row = db.session.execute(sql).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        return Resp.resp_ok(data=TestCaseObj(**row.dict()))

    async def list_test_case(self, set_id: int):
        sql = select(TestCase, TestSet, Product).outerjoin(TestSet, TestCase.set_id == TestSet.id).outerjoin(Product, TestSet.product_id == Product.id)
        sql = sql.where(TestCase.set_id == set_id)
        sql = sql.order_by(TestCase.id)
        rows: list[TestCase, TestSet, Product] = db.session.execute(sql).all()
        objs = []
        for row, row_set, row_prd in rows:
            obj = TestCaseObj(**row.dict())
            if row_set:
                obj.stage = row_set.stage
            if row_prd:
                obj.product_name = row_prd.name
                obj.product_version = row_prd.full_version
            objs.append(obj)
        return Resp.resp_ok(data=Page(rows=objs))
      
    export_columns = [
        "code",
        "srs_code"
        "test_type",
        "stage",
        "function",
        "description",
        "precondition",
        "test_step",
        "expect",
        "note"
    ]

    async def export_test_cases(self, output, *args, **kwargs):
        resp = await self.list_test_case(*args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_test_case.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            for cidx, key in enumerate(self.export_columns, 1):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
        wb.save(output)
        output.seek(0)
        