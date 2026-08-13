import logging
import os
import re
from typing import List, Tuple
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from openpyxl import load_workbook

from ..obj.vobj_user import UserObj
from ..obj.tobj_role import Roles
from ..model.doc_file import DocFile
from ..model.project import Project
from ..model.prod_dhf import ProdDhf
from ..model.test_set import TestSet
from ..model.sds_doc import SdsDoc
from ..obj.vobj_product import ProductObj, TraceObj
from ..model.product import UserProd, Product
from ..model.srs_doc import SrsDoc
from ..obj.tobj_product import ProductForm
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db
from .serv_utils import new_version

logger = logging.getLogger(__name__)

# 与 serv_srs_doc.DELETED_SRS_VERSION_PREFIX 保持一致：列表已隐藏，不算用户可见的引用
DELETED_SRS_VERSION_PREFIX = "__deleted_srs__"

DOC_FILE_CATEGORY_LABELS = {
    "img_topo": "物理拓扑图",
    "img_struct": "体系结构图",
    "img_flow": "网络安全流程图",
    "img_ui": "用户界面关系图",
    "img_home": "主页面图示",
}


def _version_seq(v: str) -> int:
    m = re.search(r"(\d+)(?!.*\d)", v or "")
    return int(m.group(1)) if m else -1


def _build_prod_docfile_block_msg(product_id: int) -> str:
    rows = db.session.execute(
        select(DocFile.category, func.count(DocFile.id))
        .where(DocFile.product_id == product_id)
        .group_by(DocFile.category)
        .order_by(DocFile.category)
    ).all()
    if not rows:
        return ts("msg_prod_x_docfile")
    parts = []
    for category, count in rows:
        label = DOC_FILE_CATEGORY_LABELS.get(str(category or "").strip(), category or "未知类型")
        parts.append(f"{label} {int(count or 0)} 条")
    detail = "、".join(parts)
    template = ts("msg_prod_x_docfile_detail")
    if "%(detail)s" in template:
        return template % {"detail": detail}
    return f"{ts('msg_prod_x_docfile')} 剩余：{detail}。"


def _active_srs_doc_filter():
    return ~SrsDoc.version.like(f"{DELETED_SRS_VERSION_PREFIX}%")


def _build_prod_srsdoc_block_msg(product_id: int) -> str:
    rows = db.session.execute(
        select(SrsDoc.version, SrsDoc.folder_name)
        .where(SrsDoc.product_id == product_id)
        .where(_active_srs_doc_filter())
        .order_by(desc(SrsDoc.id))
        .limit(8)
    ).all()
    if not rows:
        return ts("msg_prod_x_srsdoc")
    parts = []
    for version, folder_name in rows:
        label = str(version or "").strip() or "(未命名版本)"
        folder = str(folder_name or "").strip()
        parts.append(f"{label}{f'（{folder}）' if folder else ''}")
    extra = db.session.execute(
        select(func.count(SrsDoc.id))
        .where(SrsDoc.product_id == product_id)
        .where(_active_srs_doc_filter())
    ).scalar() or 0
    detail = "、".join(parts)
    if extra > len(rows):
        detail = f"{detail} 等共 {int(extra)} 份"
    template = ts("msg_prod_x_srsdoc_detail")
    if "%(detail)s" in template:
        return template % {"detail": detail}
    return f"{ts('msg_prod_x_srsdoc')} 剩余：{detail}。"


class Server(object):

    async def add_product(self, op_user: UserObj, form: ProductForm):
        try:
            sql = select(func.count(Product.id)).where(Product.name == form.name, Product.full_version == form.full_version)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            user_ids = form.user_ids or []
            form.user_ids = None
            row = Product(**form.dict(exclude_none=True))
            row.id = None
            row.create_user_id = op_user.id
            if op_user.role_code == Roles.product_manager.value.code:
                user_ids = [op_user.id]
            elif not user_ids:
                # 产品管理页目前不再传 user_ids，默认关联当前操作人，避免“新增成功但列表不可见”。
                user_ids = [op_user.id]
            db.session.add(row)
            db.session.flush()
            if user_ids:
                db.session.add_all([UserProd(user_id=user_id, product_id=row.id) for user_id in user_ids])
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def duplicate_product(self, op_user: UserObj, id: int, product_id: int = None, name: str = None, full_version: str = None, dhf_ids: List[int] = None):
        try:
            from_row: Product = db.session.execute(select(Product).where(Product.id == id)).scalars().first()
            if not from_row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            if op_user.role_code == Roles.product_manager.value.code and from_row.create_user_id != op_user.id:
                return Resp.resp_err(msg=ts("msg_no_perm"))
            target_row: Product = from_row
            if product_id and product_id != from_row.id:
                target_row = db.session.execute(select(Product).where(Product.id == product_id)).scalars().first()
                if not target_row:
                    return Resp.resp_err(msg=ts("msg_obj_null"))
            target_name = str(name or "").strip() or target_row.name
            same_name = target_name == from_row.name
            all_full_versions = db.session.execute(
                select(Product.full_version).where(Product.name == target_name)
            ).scalars().all()
            all_release_versions = db.session.execute(
                select(Product.release_version).where(Product.name == target_name)
            ).scalars().all()
            existing_full_set = {v for v in all_full_versions if v}
            if full_version and str(full_version).strip():
                next_full_version = str(full_version).strip()
            elif same_name:
                next_full_version = new_version(from_row.full_version)
            else:
                valid_full = [v for v in all_full_versions if v]
                next_full_version = new_version(max(valid_full, key=_version_seq)) if valid_full else from_row.full_version
            if not next_full_version:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            if next_full_version in existing_full_set:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            if same_name:
                release_version = new_version(from_row.release_version)
            else:
                valid_release = [v for v in all_release_versions if v]
                release_version = new_version(max(valid_release, key=_version_seq)) if valid_release else from_row.release_version
            named_row = db.session.execute(
                select(Product).where(Product.name == target_name).limit(1)
            ).scalars().first()
            ref_row = named_row or target_row
            new_row = Product(
                name=target_name,
                project_id=from_row.project_id if same_name else ref_row.project_id,
                category=from_row.category,
                type_code=from_row.type_code,
                full_version=next_full_version,
                release_version=release_version,
                udi=from_row.udi,
                product_code=from_row.product_code,
                registrant=from_row.registrant,
                scope=from_row.scope,
                component=from_row.component,
                overall_desc=from_row.overall_desc,
                note=from_row.note,
                create_user_id=op_user.id,
            )
            db.session.add(new_row)
            db.session.flush()
            user_ids = db.session.execute(
                select(UserProd.user_id).where(UserProd.product_id == from_row.id)
            ).scalars().all()
            user_ids = list(user_ids) if user_ids else []
            if op_user.role_code == Roles.product_manager.value.code:
                user_ids = [op_user.id]
            elif not user_ids:
                user_ids = [op_user.id]
            if user_ids:
                db.session.add_all([UserProd(user_id=user_id, product_id=new_row.id) for user_id in user_ids])

            from .serv_prod_runtime_env import copy_prod_runtime_env_for_product
            runtime_env_copied = copy_prod_runtime_env_for_product(from_row.id, new_row.id)

            selected_dhf_rows = []
            if dhf_ids:
                from .serv_product_dhf_copy import copy_dhf_linked_assets, apply_product_version_token
                selected_dhf_rows = db.session.execute(
                    select(ProdDhf).where(ProdDhf.id.in_(dhf_ids), ProdDhf.prod_id == from_row.id).order_by(ProdDhf.code)
                ).scalars().all()
                for src in selected_dhf_rows:
                    code = str(src.code or "").strip()
                    if not code:
                        continue
                    if same_name:
                        code = apply_product_version_token(code, next_full_version)
                    db.session.add(ProdDhf(prod_id=new_row.id, code=code, name=str(src.name or "").strip()))

            db.session.commit()

            copy_stats = {}
            if selected_dhf_rows:
                copy_stats = await copy_dhf_linked_assets(
                    from_row.id,
                    new_row.id,
                    selected_dhf_rows,
                    same_name=same_name,
                    target_full_version=next_full_version,
                )

            data = ProductForm(id=new_row.id)
            resp_fields = {}
            if runtime_env_copied:
                resp_fields["runtime_env_copied"] = 1
            if copy_stats:
                resp_fields.update({k: v for k, v in copy_stats.items() if k in ("dhf_count", "doc_count", "test_set_count", "doc_file_count")})
            if resp_fields:
                data = ProductForm(id=new_row.id, **resp_fields)
            return Resp.resp_ok(data=data)
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_product(self, id):
        sql = select(func.count(SrsDoc.id)).where(SrsDoc.product_id == id).where(_active_srs_doc_filter())
        count = db.session.execute(sql).scalar()
        if count > 0:
            return Resp.resp_err(msg=_build_prod_srsdoc_block_msg(id))

        sql = select(func.count(TestSet.id)).where(TestSet.product_id == id)
        count = db.session.execute(sql).scalar()
        if count > 0:
            return Resp.resp_err(msg=ts("msg_prod_x_testset"))
        
        sql = select(func.count(ProdDhf.id)).where(ProdDhf.prod_id == id)
        count = db.session.execute(sql).scalar()
        if count > 0:
            return Resp.resp_err(msg=ts("msg_prod_x_proddhf"))
        
        sql = select(func.count(DocFile.id)).where(DocFile.product_id == id)
        count = db.session.execute(sql).scalar()
        if count > 0:
            return Resp.resp_err(msg=_build_prod_docfile_block_msg(id))

        db.session.execute(delete(Product).where(Product.id == id))
        db.session.commit()
        return Resp.resp_ok()
   
    async def update_product(self, op_user: UserObj, form: ProductForm):
        try:
            sql = select(func.count(Product.id)).where(Product.name == form.name, Product.full_version == form.full_version, Product.id != form.id)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            sql = select(Product).where(Product.id == form.id)
            row:Product = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            if op_user.role_code == Roles.product_manager.value.code and row.create_user_id != op_user.id:
                return Resp.resp_err(msg=ts("msg_no_perm"))
            user_ids = form.user_ids
            form.user_ids = None
            if op_user.role_code == Roles.product_manager.value.code:
                user_ids = [op_user.id]
            for key, value in form.dict(exclude_none=True).items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            if user_ids is not None:
                db.session.execute(delete(UserProd).where(UserProd.product_id == row.id))
                if user_ids:
                    db.session.add_all([UserProd(user_id=user_id, product_id=row.id) for user_id in user_ids])
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def get_product(self, id:str, with_trace: int = 0):
        sql = select(Product).where(Product.id == id)
        row = db.session.execute(sql).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        obj = ProductObj(**row.dict())
        if row.project_id:
            proj = db.session.execute(select(Project).where(Project.id == row.project_id)).scalars().first()
            if proj:
                obj.country = proj.country
        if with_trace:
            traces_dict = self.__query_traces([row])
            obj.traces = traces_dict.get(row.id, [])
            obj.srs_versions = list(dict.fromkeys([trace.srsdoc_version for trace in obj.traces]))
            obj.sds_versions = list(dict.fromkeys([trace.sdsdoc_version for trace in obj.traces]))
        return Resp.resp_ok(data=obj)
    
    async def export_product_trace(self, output, id: int):
        def __fix(rid, prod: ProductObj):
            ws.cell(row=rid, column=1, value=prod.name)
            ws.cell(row=rid, column=2, value=prod.type_code)
            ws.cell(row=rid, column=3, value=prod.full_version)

        resp = await self.get_product(id, with_trace=1)
        prod = resp.data
        if not prod:
            return
        temp_path = os.path.join(os.path.dirname(__file__), "temp_product_trace.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        rid = 1
        for srs in prod.srs_versions:
            rid += 1
            __fix(rid, prod)
            ws.cell(row=rid, column=4, value=ts("product.doc_srs"))
            ws.cell(row=rid, column=5, value=srs)
        for sds in prod.sds_versions:
            rid += 1
            __fix(rid, prod)
            ws.cell(row=rid, column=4, value=ts("product.doc_sds"))
            ws.cell(row=rid, column=5, value=sds)
        wb.save(output)
        output.seek(0)

    def __query_traces(self, objs: List[Product]):
        sql = select(SdsDoc, SrsDoc).join(SrsDoc, SdsDoc.srsdoc_id == SrsDoc.id).where(SrsDoc.product_id.in_([obj.id for obj in objs])).order_by(SrsDoc.id, SdsDoc.id)
        rows: List[SdsDoc, SrsDoc] = db.session.execute(sql).all()
        result_dict = dict()
        for row_sds, row_srs in rows:
            trace = TraceObj(sdsdoc_version=row_sds.version, srsdoc_version=row_srs.version)
            result_dict.setdefault(row_srs.product_id, []).append(trace)
        return result_dict

    async def list_product(self, op_user: UserObj, export = False, fuzzy: str = None, with_trace: int = 0, page_index: int = 0, page_size: int = 10):
        def __query_users(prod_ids: List[int]):
            result_dict = dict()
            if prod_ids:
                sql = select(UserProd.user_id, UserProd.product_id).where(UserProd.product_id.in_(prod_ids))
                rows: List[Tuple[int, int]] = db.session.execute(sql).all()
                for row_user, row_prod in rows:
                    result_dict.setdefault(row_prod, []).append(row_user)
            return result_dict
        
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 
    
        sql = select(Product, Project).outerjoin(Project, Product.project_id == Project.id)
        if fuzzy:
            sql = sql.where(
                or_(
                    Product.name.like(f"%{fuzzy}%"),
                    Product.category.like(f"%{fuzzy}%"),
                    Product.type_code.like(f"%{fuzzy}%"),
                    Product.full_version.like(f"%{fuzzy}%"),
                    Product.release_version.like(f"%{fuzzy}%"),
                    Product.udi.like(f"%{fuzzy}%"),
                    Product.product_code.like(f"%{fuzzy}%"),
                    Product.scope.like(f"%{fuzzy}%"),
                    Product.component.like(f"%{fuzzy}%"),
                    Product.note.like(f"%{fuzzy}%"),
                )
            )
        # 数据可见范围：产品经理只看自己创建的产品；超管及其它角色（DQA/RA/QA/开发/测试）查看全部产品
        if op_user.id != 1 and op_user.role_code == Roles.product_manager.value.code:
            sql = sql.where(Product.create_user_id == op_user.id)
        
        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.offset(page_size * page_index).limit(page_size).order_by(desc(Product.create_time))
        rows: List[Tuple[Product, Project]] = db.session.execute(sql).all()
        prod_users = __query_users( [row[0].id for row in rows])
        objs: List[ProductObj] = []
        for row, row_proj in rows:
            obj = ProductObj(**row.dict())
            obj.user_ids = prod_users.get(row.id, [])
            if row_proj:
                obj.country = row_proj.country
            objs.append(obj)
        if with_trace:
            traces_dict = self.__query_traces(objs)
            for obj in objs:
                obj.traces = traces_dict.get(obj.id, [])
                obj.srs_versions = list(dict.fromkeys([trace.srsdoc_version for trace in obj.traces]))
                obj.sds_versions = list(dict.fromkeys([trace.sdsdoc_version for trace in obj.traces]))
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))
      
    export_columns = [
        "name",
        "country",
        "category",
        "type_code",
        "full_version",
        "release_version",
        "udi",
        "product_code",
        "scope",
        "component",
        "note",
        "create_time"
    ]

    async def export_products(self, output, op_user: UserObj, *args, **kwargs):
        resp = await self.list_product(op_user, export=True, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_product.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            for cidx, key in enumerate(self.export_columns, 1):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
        wb.save(output)
        output.seek(0)
        