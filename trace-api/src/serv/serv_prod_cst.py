import logging
import os
from typing import List
from sqlalchemy import select, delete, func
from sqlalchemy.sql import desc
from sqlalchemy.dialects.postgresql import insert as pg_insert
from openpyxl import load_workbook
from ..obj.vobj_user import UserObj
from ..model.product import Product, UserProd
from ..model.cst import Cst
from ..model.cst_rcm import CstRcm
from ..model.rcm import Rcm
from ..model.prod_cst import ProdCst
from ..obj.tobj_prod_cst import ProdCstForm, ProdCstsForm
from ..obj.vobj_prod_cst import ProdCstObj
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):

    async def add_prod_csts(self, form: ProdCstsForm):
        try:
            cst_ids = form.cst_ids or []
            if not cst_ids:
                return Resp.resp_ok()
            # 从 CST 总表(cst_rcm)带入每个威胁已维护的关联 RCM 编号，写入新行的 rcm_codes（逗号分隔）
            rcm_map = {}
            sql = (
                select(CstRcm.cst_id, Rcm.code)
                .join(Rcm, CstRcm.rcm_id == Rcm.id)
                .where(CstRcm.cst_id.in_(list(cst_ids)))
                .order_by(Rcm.code)
            )
            for cst_id, code in db.session.execute(sql).all():
                rcm_map.setdefault(cst_id, []).append(code)
            values = []
            for cst_id in cst_ids:
                codes = rcm_map.get(cst_id) or []
                values.append(dict(prod_id=form.prod_id, cst_id=cst_id, rcm_codes=",".join(codes) if codes else None))
            db.session.execute(pg_insert(ProdCst).values(values).on_conflict_do_nothing())
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
    
    async def update_prod_cst(self, form: ProdCstForm):
        try:
            sql = select(ProdCst).where(ProdCst.id == form.id)
            row:ProdCst = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            update_keys = set(["prev_score", "prev_severity", "prev_level", "prev_accept", "cur_score", "cur_severity", "cur_level", "cur_accept", "rcm_codes"])
            for key, value in form.dict().items():
                if key not in update_keys or value is None:
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_prod_csts(self, ids: List[str]):
        if ids:
            db.session.execute(delete(ProdCst).where(ProdCst.id.in_(ids)))
            db.session.commit()
        return Resp.resp_ok()

    async def list_prod_cst(self, op_user: UserObj, export = False, prod_id: int = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 

        sql = select(ProdCst, Product, Cst).outerjoin(Product, ProdCst.prod_id == Product.id).outerjoin(Cst, ProdCst.cst_id == Cst.id)
        if prod_id:
            sql = sql.where(ProdCst.prod_id == prod_id)
        if not prod_id and op_user and op_user.id != 1:
            subquery = select(UserProd.product_id).where(UserProd.user_id == op_user.id).scalar_subquery()
            sql = sql.where(Product.id.in_(subquery))
        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.order_by(Cst.code)
        rows: List[ProdCst, Product, Cst] = db.session.execute(sql).all()
        objs = []
        for row, row_prd, row_cst in rows:
            obj = ProdCstObj(**row.dict())
            if row_cst:
                obj.code = row_cst.code
                obj.category = row_cst.category
                obj.module = row_cst.module
                obj.connection = row_cst.connection
                obj.description = row_cst.description
                obj.harm = row_cst.harm
            if row_prd:
                obj.product_name = row_prd.name
                obj.product_version = row_prd.full_version
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))

    export_columns = [
        "code",
        "category",
        "description",

        "prev_score",
        "prev_severity",
        "prev_level",
        "prev_accept",

        "cur_score",
        "cur_severity",
        "cur_level",
        "cur_accept",

        "rcm_codes",
    ]

    async def export_prod_csts(self, op_user: UserObj, output, *args, **kwargs):
        resp = await self.list_prod_cst(op_user, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_prod_cst.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            for cidx, key in enumerate(self.export_columns, 1):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
        wb.save(output)
        output.seek(0)
        