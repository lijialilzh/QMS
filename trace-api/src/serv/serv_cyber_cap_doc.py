#!/usr/bin/env python
# encoding: utf-8

# 网络安全能力分析（MDS2）服务层。
# 文档可编辑部分（22 项安全能力 + 健康数据管理的 是/否/不适用 答案与各项备注）以
#   content(JSON: {"cells": {单元格坐标: 值}}) 存储，覆盖到 xlsx 模板对应单元格。
# 产品相关内容（公司信息/预期用途/运行环境/器械型号/软件修订版/文件号/日期）导出时自动填充：
#   - 公司信息  ← company_info（注册人/住所/联系电话/代表人）
#   - 预期用途  ← product.scope（试用范围）
#   - 运行环境  ← prod_runtime_env（服务器/用户端操作系统、浏览器）填入 RDMP 备注
#   - 器械型号  ← product.type_code
#   - 软件修订版 ← product.full_version
#   - 文件ID    ← 产品 DHF（按文档名匹配）
#   - 文件/软件发布日期 ← 项目时间逻辑线（关键字匹配的最新日期）

import copy
import json
import logging
import os
import re
from typing import List
from sqlalchemy import delete, func, select

from openpyxl import load_workbook

from ..model.product import Product
from ..model.prod_dhf import ProdDhf
from ..model.company_info import CompanyInfo
from ..model.prod_runtime_env import ProdRuntimeEnv
from ..model.cyber_cap_doc import CyberCapDoc
from ..model.project_timeline import ProjectTimelineRow, ProjectTimelineCell
from ..obj import Page, Resp
from ..obj.tobj_role import Roles
from ..obj.vobj_user import UserObj
from ..obj.tobj_cyber_cap_doc import CyberCapDocForm
from ..obj.vobj_cyber_cap_doc import CyberCapDocObj
from ..utils.i18n import ts
from ..utils.sql_ctx import db
from . import msg_err_db
from . import serv_review_util
from .serv_utils import new_version, sync_file_no_version
from .serv_prod_runtime_env import DEFAULT_RUNTIME_ENV

logger = logging.getLogger(__name__)

# 本文档名（用于从产品 DHF 匹配文件编号）
DOC_NAME = "网络安全能力分析"
# 文件/软件发布日期从时间逻辑线匹配的关键字（取命中行最新日期）
DATE_KEYWORDS = ["网络安全能力分析", "网络安全"]
# RDMP 备注第 2 条为固定说明（运行环境只覆盖第 1 条 OS 清单）
RDMP_NOTE2 = "2.《自研软件网络安全报告》附录B包含本产品全部现成软件的清单，第2.5章公布了维护计划"
SHEET_NAME = "MDS2 2013"

_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "cyber_mds2_template.xlsx")
_SCHEMA_PATH = os.path.join(os.path.dirname(__file__), "cyber_mds2_schema.json")

# 头部重复字段（标签下方单元格写值）
_LABEL_PROD_FIELDS = {
    "制造商": "registrant",
    "公司名称": "registrant",
    "文件ID": "file_no",
    "器械型号": "type_code",
    "软件修订版": "full_version",
    "文件发布日期": "date",
    "软件发布日期": "date",
}
_LABEL_DATE = {"文件发布日期", "软件发布日期"}


def _load_schema():
    try:
        with open(_SCHEMA_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        logger.exception("load cyber_mds2_schema.json failed")
        return {"sections": [], "rdmp_cells": []}


SCHEMA = _load_schema()


class Server(object):

    # ---------------- 规范化 ----------------
    def __normalize_content(self, content):
        cells = {}
        if isinstance(content, dict) and isinstance(content.get("cells"), dict):
            for k, v in content["cells"].items():
                cells[str(k)] = "" if v is None else str(v)
        return {"cells": cells}

    # ---------------- 自动获取 ----------------
    def __collect_autofill(self, prod_id, product, doc_version):
        def to_int(v):
            digits = re.sub(r"[^\d]", "", str(v or ""))
            return int(digits) if digits else None

        # 公司信息：优先按产品注册人匹配，未匹配再取首条
        prod_registrant = (getattr(product, "registrant", "") or "").strip() if product else ""
        company = None
        if prod_registrant:
            company = db.session.execute(
                select(CompanyInfo).where(CompanyInfo.registrant == prod_registrant)
            ).scalars().first()
        if not company:
            company = db.session.execute(select(CompanyInfo).order_by(CompanyInfo.id.asc())).scalars().first()
        registrant = (getattr(company, "registrant", "") or "").strip() if company else prod_registrant
        address = (getattr(company, "address", "") or "").strip() if company else ""
        phone = (getattr(company, "contact_phone", "") or "").strip() if company else ""
        representative = (getattr(company, "representative", "") or "").strip() if company else ""

        # 运行环境：无记录时回退运行环境模板默认值（与运行环境页面预填一致）
        env = db.session.execute(select(ProdRuntimeEnv).where(ProdRuntimeEnv.prod_id == prod_id)).scalars().first()

        def env_val(attr):
            v = (getattr(env, attr, "") or "").strip() if env else ""
            return v or (DEFAULT_RUNTIME_ENV.get(attr) or "")

        srv_os = env_val("srv_os")
        cli_os = env_val("cli_os")
        cli_browser = env_val("cli_browser")
        runtime = ""
        if srv_os or cli_os or cli_browser:
            l1 = f"1.服务器端：{srv_os}"
            l2 = f"客户端：{cli_os}"
            if cli_browser:
                l2 += f"；{cli_browser}"
            runtime = f"{l1}\n{l2}\n{RDMP_NOTE2}"

        # 时间逻辑线日期
        tl_rows = db.session.execute(
            select(ProjectTimelineRow).where(ProjectTimelineRow.prod_id == prod_id)
        ).scalars().all()
        cell_map = {}
        if tl_rows:
            for c in db.session.execute(
                select(ProjectTimelineCell).where(ProjectTimelineCell.row_id.in_([r.id for r in tl_rows]))
            ).scalars().all():
                cell_map.setdefault(c.row_id, []).append(c.output_result or "")
        date_rows = [r for r in tl_rows if (r.row_type or "date") == "date" and to_int(r.year) and to_int(r.month)]

        def date_key(r):
            return to_int(r.year) * 10000 + to_int(r.month) * 100 + (to_int(r.day) or 0)

        date = ""
        hit = [r for r in date_rows if any(any(k in str(v or "") for k in DATE_KEYWORDS) for v in cell_map.get(r.id, []))]
        if hit:
            r = max(hit, key=date_key)
            date = f"{to_int(r.year)}年{to_int(r.month)}月{to_int(r.day)}日"

        return {
            "registrant": registrant,
            "address": address,
            "phone": phone,
            "representative": representative,
            "scope": (getattr(product, "scope", "") or "").strip() if product else "",
            "type_code": (getattr(product, "type_code", "") or "").strip() if product else "",
            "full_version": (getattr(product, "full_version", "") or "").strip() if product else "",
            "file_no": self.__dhf_file_no(prod_id),
            "date": date,
            "runtime": runtime,
        }

    def __dhf_file_no(self, prod_id):
        row = db.session.execute(
            select(ProdDhf).where(ProdDhf.prod_id == prod_id, ProdDhf.name.like(f"%{DOC_NAME}%"))
        ).scalars().first()
        return (row.code or "").strip() if row and row.code else ""

    def __to_obj(self, row: CyberCapDoc, product: Product = None):
        obj = CyberCapDocObj(**row.dict())
        obj.content = self.__normalize_content(obj.content)
        if product:
            obj.product_name = product.name
            obj.product_version = product.full_version
            obj.product_full_version = product.full_version
            obj.product_type_code = product.type_code
            info = self.__collect_autofill(product.id, product, row.version)
            # 文件编号优先用文档已填值，未填时回退 DHF
            if not (obj.file_no or "").strip():
                resolved = serv_review_util.resolve_doc_file_no(product.id, obj.file_no, row.version, "cyber_cap")
                if resolved:
                    obj.file_no = resolved
                elif info.get("file_no"):
                    obj.file_no = info["file_no"]
            # auto 预览里的文件编号同步为最终展示值
            info["file_no"] = obj.file_no or info.get("file_no") or ""
            obj.auto = info
        return obj

    # ---------------- CRUD ----------------
    async def add_cyber_cap_doc(self, form: CyberCapDocForm):
        try:
            sql = select(func.count(CyberCapDoc.id)).where(
                CyberCapDoc.product_id == form.product_id,
                CyberCapDoc.version == form.version,
            )
            if db.session.execute(sql).scalar() > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            row = CyberCapDoc(**form.dict(exclude_none=True))
            row.id = None
            row.file_no = serv_review_util.resolve_doc_file_no(form.product_id, form.file_no, form.version, "cyber_cap") or None
            row.content = self.__normalize_content(row.content)
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok(data=CyberCapDocForm(id=row.id))
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def duplicate_cyber_cap_doc(self, id: int, product_id: int = None):
        try:
            fromdoc: CyberCapDoc = db.session.execute(select(CyberCapDoc).where(CyberCapDoc.id == id)).scalars().first()
            if not fromdoc:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            target_pid = product_id or fromdoc.product_id
            all_versions = db.session.execute(select(CyberCapDoc.version).where(CyberCapDoc.product_id == target_pid)).scalars().all()
            existing_set = {v for v in all_versions if v}
            if target_pid == fromdoc.product_id:
                version = new_version(fromdoc.version)
            else:
                def _seq(v):
                    m = re.search(r"(\d+)(?!.*\d)", v or "")
                    return int(m.group(1)) if m else -1
                valid = [v for v in all_versions if v]
                version = new_version(max(valid, key=_seq)) if valid else fromdoc.version
            while version in existing_set:
                version = new_version(version)
            newdoc = CyberCapDoc(
                product_id=target_pid,
                version=version,
                file_no=sync_file_no_version((fromdoc.file_no or "").strip() or self.__dhf_file_no(target_pid), version) or None,
                change_log=fromdoc.change_log,
                content=copy.deepcopy(self.__normalize_content(fromdoc.content)),
            )
            db.session.add(newdoc)
            db.session.commit()
            return Resp.resp_ok(data=CyberCapDocForm(id=newdoc.id))
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def update_cyber_cap_doc(self, form: CyberCapDocForm):
        try:
            row: CyberCapDoc = db.session.execute(select(CyberCapDoc).where(CyberCapDoc.id == form.id)).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict(exclude_none=True).items():
                if key == "id":
                    continue
                if key == "content":
                    value = self.__normalize_content(value)
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def delete_cyber_cap_doc(self, id: int):
        db.session.execute(delete(CyberCapDoc).where(CyberCapDoc.id == id))
        db.session.commit()
        return Resp.resp_ok()

    async def get_cyber_cap_doc(self, id: int):
        sql = select(CyberCapDoc, Product).join(Product, CyberCapDoc.product_id == Product.id).where(CyberCapDoc.id == id)
        row = db.session.execute(sql).first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        doc, product = row
        return Resp.resp_ok(data=self.__to_obj(doc, product))

    async def list_cyber_cap_doc(self, op_user: UserObj = None, product_id: int = 0, version: str = None, page_index: int = 0, page_size: int = 10):
        wheres = []
        if product_id:
            wheres.append(CyberCapDoc.product_id == product_id)
        if version:
            wheres.append(CyberCapDoc.version.like(f"%{version}%"))
        if op_user and op_user.id != 1 and op_user.role_code == Roles.product_manager.value.code:
            wheres.append(Product.create_user_id == op_user.id)
        sql_total = select(func.count(CyberCapDoc.id)).join(Product, CyberCapDoc.product_id == Product.id).where(*wheres)
        total = db.session.execute(sql_total).scalar() or 0
        sql = (
            select(CyberCapDoc, Product)
            .join(Product, CyberCapDoc.product_id == Product.id)
            .where(*wheres)
            .order_by(CyberCapDoc.id.desc())
            .offset(page_index * page_size)
            .limit(page_size)
        )
        rows: List[CyberCapDocObj] = [self.__to_obj(doc, product) for doc, product in db.session.execute(sql).all()]
        return Resp.resp_ok(data=Page(total=total, rows=rows, page_index=page_index, page_size=page_size))

    def get_schema(self):
        return Resp.resp_ok(data=SCHEMA)

    async def autofill_preview(self, product_id: int):
        product = db.session.execute(select(Product).where(Product.id == product_id)).scalars().first()
        if not product:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        info = self.__collect_autofill(product_id, product, "")
        info["product_name"] = product.name
        info["product_full_version"] = product.full_version
        info["product_type_code"] = product.type_code
        return Resp.resp_ok(data=info)

    # ---------------- 导出 xlsx ----------------
    async def export_cyber_cap_doc(self, output, id: int):
        resp = await self.get_cyber_cap_doc(id)
        obj: CyberCapDocObj = resp.data
        wb = load_workbook(_TEMPLATE_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

        merged_ranges = list(ws.merged_cells.ranges)

        def anchor(coord):
            for mr in merged_ranges:
                if coord in mr:
                    return str(mr.coord).split(":")[0]
            return coord

        def put(coord, val):
            try:
                ws[anchor(coord)] = "" if val is None else val
            except Exception:
                logger.exception("write cell %s failed", coord)

        if obj is not None:
            # 1) 用户可编辑单元格（答案/备注）覆盖模板
            cells = (self.__normalize_content(obj.content).get("cells") or {})
            for coord, val in cells.items():
                put(coord, val)

            # 2) 产品相关自动字段
            info = obj.auto or {}
            for row in ws.iter_rows(min_row=1, max_row=229):
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    key = str(v).strip()
                    fld = _LABEL_PROD_FIELDS.get(key)
                    if not fld:
                        continue
                    below = cell.offset(row=1, column=0).coordinate
                    val = info.get(fld) or ""
                    if key in _LABEL_DATE:
                        put(below, val)  # 日期即使为空也写，清掉占位 0
                    elif val:
                        put(below, val)
            if info.get("scope"):
                put("B15", info["scope"])
            if info.get("address") or info.get("phone"):
                put("G11", f"地址：{info.get('address') or ''}\n电话：{info.get('phone') or ''}")
            if info.get("representative"):
                put("E13", info["representative"])
            if info.get("runtime"):
                put("D190", info["runtime"])

        wb.save(output)
        output.seek(0)
