import io
import logging
import re
from openpyxl import load_workbook
from sqlalchemy import select, delete, func
from sqlalchemy.sql import asc
from ..model.project_member import ProjectMember
from ..obj.tobj_project_member import ProjectMemberForm
from ..obj.vobj_project_member import ProjectMemberObj
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):

    async def add_project_member(self, form: ProjectMemberForm):
        try:
            if not form.prod_id:
                return Resp.resp_err(msg=ts("msg_err_param"))
            row = ProjectMember(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def update_project_member(self, form: ProjectMemberForm):
        try:
            sql = select(ProjectMember).where(ProjectMember.id == form.id)
            row: ProjectMember = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key in ("id", "prod_id"):
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def delete_project_members(self, ids: list):
        ids = [int(i) for i in (ids or []) if str(i).strip().isdigit()]
        if not ids:
            return Resp.resp_err(msg=ts("msg_err_param"))
        db.session.execute(delete(ProjectMember).where(ProjectMember.id.in_(ids)))
        db.session.commit()
        return Resp.resp_ok()

    async def list_project_member(self, prod_id: int = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10

        sql = select(ProjectMember)
        if prod_id:
            sql = sql.where(ProjectMember.prod_id == prod_id)

        sql_count = select(func.count()).select_from(sql)
        total = db.session.execute(sql_count).scalars().first()

        sql = sql.order_by(asc(ProjectMember.sort_order), asc(ProjectMember.id))
        sql = sql.offset(page_size * page_index).limit(page_size)
        rows: list[ProjectMember] = db.session.execute(sql).scalars().all()
        objs = [ProjectMemberObj(**row.dict()) for row in rows]
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))

    async def import_project_members(self, prod_id: int, file_bytes: bytes, replace: bool = True):
        try:
            if not prod_id:
                return Resp.resp_err(msg=ts("msg_err_param"))
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            # 优先选「职能/人员」表头的 sheet，其次标题含「人员」，再退化为首个 sheet
            target_ws = None
            for ws in wb.worksheets:
                first = [str(c or "").strip() for c in (next(ws.iter_rows(values_only=True), ()) or ())]
                if "职能" in first and "人员" in first:
                    target_ws = ws
                    break
            if target_ws is None:
                for ws in wb.worksheets:
                    if "人员" in str(ws.title or ""):
                        target_ws = ws
                        break
            if target_ws is None:
                target_ws = wb.worksheets[0]

            grid = [list(r) for r in target_ws.iter_rows(values_only=True)]
            if not grid:
                return Resp.resp_err(msg=ts("msg_err_param"))
            # 表头行：定位「职能」「人员」列
            header = [str(c or "").strip() for c in grid[0]]
            role_col = header.index("职能") if "职能" in header else 0
            name_col = header.index("人员") if "人员" in header else 1

            if replace:
                db.session.execute(delete(ProjectMember).where(ProjectMember.prod_id == prod_id))

            last_role = ""
            sort_order = 0
            imported = 0
            for r in grid[1:]:
                def cell(i):
                    return str(r[i]).strip() if i < len(r) and r[i] is not None else ""
                role_raw = cell(role_col)
                name_raw = cell(name_col)
                if role_raw:
                    last_role = role_raw
                role = role_raw or last_role
                if not (role or name_raw):
                    continue
                # 单格多人（、,/ 空白 换行）拆成一人一行
                names = [n for n in re.split(r"[、,，/\s\n]+", name_raw) if n]
                if not names:
                    names = [""]
                for nm in names:
                    sort_order += 1
                    db.session.add(ProjectMember(prod_id=prod_id, role=role, name=nm, sort_order=sort_order))
                    imported += 1
            db.session.commit()
            return Resp.resp_ok(data={"imported": imported})
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
