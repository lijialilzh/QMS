import logging
import io
import os
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from ..model.prod_cst import ProdCst
from ..obj.vobj_cst import CstObj
from ..model.cst import Cst
from ..obj.tobj_cst import CstForm
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):
    @staticmethod
    def __to_str(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def __read_cell_value(cell, ws_raw):
        value = cell.value
        if value is not None and str(value).strip() != "":
            return value
        raw = ws_raw.cell(row=cell.row, column=cell.column).value
        if isinstance(raw, str) and raw.startswith("="):
            return raw
        return value

    async def add_cst(self, form: CstForm):
        try:
            sql = select(func.count(Cst.id)).where(Cst.code == form.code)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            
            row = Cst(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def import_csts(self, file):
        try:
            bys = await file.read()
            wb = load_workbook(io.BytesIO(bys), data_only=True)
            wb_raw = load_workbook(io.BytesIO(bys), data_only=False)
            ws = wb[wb.sheetnames[0]]
            ws_raw = wb_raw[wb_raw.sheetnames[0]]
            affected = 0
            for row in ws.iter_rows(min_row=2):
                vals = [self.__read_cell_value(cell, ws_raw) for cell in row[:len(self.export_columns)]]
                if not vals or all((v is None or str(v).strip() == "") for v in vals):
                    continue
                code = self.__to_str(vals[0])
                if not code:
                    continue
                data = dict(
                    code=code,
                    category=self.__to_str(vals[1]),
                    module=self.__to_str(vals[2]),
                    connection=self.__to_str(vals[3]),
                    description=self.__to_str(vals[4]),
                    harm=self.__to_str(vals[5]),
                )
                existed = db.session.execute(select(Cst).where(Cst.code == code)).scalars().first()
                if existed:
                    for k, v in data.items():
                        setattr(existed, k, v)
                else:
                    db.session.add(Cst(**data))
                affected += 1
            db.session.commit()
            return Resp.resp_ok(data={"count": affected})
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_cst(self, id):
        db.session.execute(delete(Cst).where(Cst.id == id))
        db.session.execute(delete(ProdCst).where(ProdCst.cst_id == id))
        db.session.commit()
        return Resp.resp_ok()
   
    async def update_cst(self, form: CstForm):
        try:
            sql = select(Cst).where(Cst.id == form.id)
            row:Cst = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def get_cst(self, id:str):
        sql = select(Cst).where(Cst.id == id)
        row = db.session.execute(sql).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        return Resp.resp_ok(data=CstObj(**row.dict()))

    async def list_cst(self, export = False, fuzzy: str = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 
    
        sql = select(Cst)
        if fuzzy:
            sql = sql.where(
                or_(
                    Cst.code.like(f"%{fuzzy}%"),
                    Cst.module.like(f"%{fuzzy}%"),
                    Cst.connection.like(f"%{fuzzy}%"),
                    Cst.description.like(f"%{fuzzy}%"),
                    Cst.harm.like(f"%{fuzzy}%"),
                )
            )
        
        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.offset(page_size * page_index).limit(page_size).order_by(Cst.code)
        rows: list[Cst] = db.session.execute(sql).scalars().all()
        objs = []
        for row in rows:
            obj = CstObj(**row.dict())
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))
      
    export_columns = [
        "code",
        "category",
        "module",
        "connection",
        "description",
        "harm",
        "create_time"
    ]

    async def export_csts(self, output, *args, **kwargs):
        resp = await self.list_cst(export=True, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_cst.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            for cidx, key in enumerate(self.export_columns, 1):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
                ws.cell(row=ridx, column=cidx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        wb.save(output)
        output.seek(0)
