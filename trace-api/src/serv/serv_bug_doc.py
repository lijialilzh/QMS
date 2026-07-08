#!/usr/bin/env python
# encoding: utf-8

# Bug管理及回归测试服务层（测试文件）：上传原始 xlsx 只读存档，解析「bug描述」页得到缺陷统计。

import io
import logging
import os
import re
import uuid
import zipfile
from typing import List
from xml.sax.saxutils import escape

from sqlalchemy import delete, func, select
from openpyxl import load_workbook
from starlette.concurrency import run_in_threadpool

from . import serv_review_util as ru

from ..model.product import Product
from ..model.bug_doc import BugDoc
from ..model.prod_dhf import ProdDhf
from ..obj import Page, Resp
from ..obj.tobj_role import Roles
from ..obj.vobj_user import UserObj
from ..obj.tobj_bug_doc import BugDocForm
from ..obj.vobj_bug_doc import BugDocObj
from ..utils.i18n import ts
from ..utils.sql_ctx import db
from . import msg_err_db
from .serv_utils import new_version, sync_file_no_version

logger = logging.getLogger(__name__)

DOC_NAME = "Bug管理及回归测试"
_BASE = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
TEMPLATE_PATH = os.path.join(_BASE, "src-res", "bug_template.xlsx")
DATA_DIR = os.path.join(_BASE, "data.trace", "bug_docs")
LEVELS = ["一级", "二级", "三级", "四级", "五级"]
RESOLVED_KW = ("已关闭", "已解决", "关闭", "解决", "已修复", "修复", "closed", "resolved", "fixed")


class Server(object):

    def __save_to_disk(self, data, filename):
        os.makedirs(DATA_DIR, exist_ok=True)
        ext = os.path.splitext(filename or "")[1] or ".xlsx"
        path = os.path.join(DATA_DIR, uuid.uuid4().hex + ext)
        with open(path, "wb") as f:
            f.write(data)
        return path

    def __ensure_disk(self, row):
        """返回可用磁盘路径：优先 file_path；旧数据在 file_data 里则迁移到磁盘。"""
        if row.file_path and os.path.exists(row.file_path):
            return row.file_path
        if row.file_data:
            try:
                path = self.__save_to_disk(row.file_data, row.file_name or "bug.xlsx")
                row.file_path = path
                row.file_data = None
                db.session.commit()
                return path
            except Exception:
                logger.exception("migrate bug blob to disk failed")
                db.session.rollback()
        return None

    @staticmethod
    def __sign_png(name):
        """按姓名取签名图 PNG 字节（来自人员签名管理，data URL）。取不到返回 None。"""
        url = ru._sign_by_name(name) or ""
        if not url.startswith("data:image"):
            return None
        try:
            import base64
            b64 = url.split(",", 1)[1]
            return base64.b64decode(b64)
        except Exception:
            return None

    def __cover_plan(self, row):
        """计算封面回填方案。返回 (texts, imgs, sign_urls)：
          - texts: {坐标: 文字}（日期/生效日期/修订记录；签名者无签名图时的姓名兜底）
          - imgs:  [(坐标, png字节)]（编制/审核/批准人有签名图时插图）
          - sign_urls: {坐标: data_url}（供在线预览显示签名章）
        纯 DB 读取，不做文件 IO。"""
        prod_id = row.product_id
        rev_date = ru.cover_date(prod_id, "bug")
        names = ru.cover_signer_names(prod_id, "bug", rev_date)
        drafter = names.get("编制人", "")
        reviewer = names.get("审核人", "")
        approver = names.get("批准人", "")
        ver = ((row.file_no or self.__dhf_file_no(prod_id) or "").rsplit("-", 1)[-1] or "A0").strip() or "A0"
        texts = {
            "G7": rev_date, "G8": rev_date, "G9": rev_date,   # 编制/审核/批准 日期
            "E10": rev_date,                                  # 生效日期
            "C31": rev_date, "D31": ver, "E31": "首次发布",    # 文件修订记录首行
            "G31": drafter, "H31": approver,                  # 修订人 / 批准人
        }
        imgs = []
        sign_urls = {}
        for coord, name in (("E7", drafter), ("E8", reviewer), ("E9", approver)):
            png = self.__sign_png(name)
            if png:
                imgs.append((coord, png))
                sign_urls[coord] = ru._sign_by_name(name)
            elif name:
                texts[coord] = name  # 无签名图时回退姓名文字
        return {k: v for k, v in texts.items() if v}, imgs, sign_urls

    @staticmethod
    def __set_cell(xml, coord, value):
        """把封面页里坐标为 coord 的空单元格 <c r=".." s=".."/> 就地改成内联字符串文本。
        只替换该单元格，其余 XML（含 WPS 截图、其它页）原样保留。"""
        val = escape(str(value))
        pat = re.compile(r'<c r="' + re.escape(coord) + r'"([^>/]*)/>')

        def _repl(m):
            attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
            return ('<c r="' + coord + '"' + attrs + ' t="inlineStr">'
                    '<is><t xml:space="preserve">' + val + '</t></is></c>')

        new, n = pat.subn(_repl, xml)
        if n:
            return new
        # 兼容非自闭合单元格：替换整段 <c ...>...</c>
        pat2 = re.compile(r'<c r="' + re.escape(coord) + r'"([^>]*)>.*?</c>', re.S)

        def _repl2(m):
            attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
            return ('<c r="' + coord + '"' + attrs + ' t="inlineStr">'
                    '<is><t xml:space="preserve">' + val + '</t></is></c>')

        new, _ = pat2.subn(_repl2, xml)
        return new

    def __ensure_filled(self, src_path, values, imgs):
        """返回封面已回填的 xlsx 磁盘路径（带缓存）。cover 内容不变则复用缓存，不再重复生成。
        纯文件/CPU 操作，可放线程池执行；出错回退原文件路径。"""
        import hashlib
        import json
        img_sig = [(c, hashlib.md5(p).hexdigest()) for c, p in imgs]
        raw = json.dumps({"t": values, "i": img_sig}, ensure_ascii=False, sort_keys=True)
        key = hashlib.md5(raw.encode("utf-8")).hexdigest()
        filled = src_path + ".filled.xlsx"
        keyf = src_path + ".filled.key"
        try:
            if os.path.exists(filled) and os.path.exists(keyf):
                with open(keyf, "r", encoding="utf-8") as f:
                    if f.read().strip() == key:
                        return filled
            data = self.__build_filled(src_path, values, imgs)
            with open(filled, "wb") as f:
                f.write(data)
            with open(keyf, "w", encoding="utf-8") as f:
                f.write(key)
            return filled
        except Exception:
            logger.exception("ensure filled failed")
            return src_path

    @staticmethod
    def __png_size(data):
        """从 PNG 字节读出 (宽, 高) 像素；失败返回 (200, 80)。"""
        try:
            if data[:8] == b"\x89PNG\r\n\x1a\n":
                w = int.from_bytes(data[16:20], "big")
                h = int.from_bytes(data[20:24], "big")
                if w and h:
                    return w, h
        except Exception:
            pass
        return 200, 80

    def __drawing_xml(self, imgs):
        """为签名图生成 xl/drawings 的 wsDr XML 及其 rels、媒体清单。
        返回 (drawing_xml, drawing_rels_xml, media[(zip名, 字节)], rid_used)。"""
        NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        EMU_PX = 9525
        target_h_px = 52  # 签名章目标高度（像素），宽按原图比例
        anchors = []
        rels = []
        media = []
        for i, (coord, png) in enumerate(imgs, start=1):
            col = "".join(ch for ch in coord if ch.isalpha())
            rn = int("".join(ch for ch in coord if ch.isdigit()))
            ci = 0
            for ch in col:
                ci = ci * 26 + (ord(ch.upper()) - 64)
            ci -= 1
            ri = rn - 1
            w_px, h_px = self.__png_size(png)
            cy = target_h_px * EMU_PX
            cx = int(cy * (w_px / float(h_px)))
            rid = "rId%d" % i
            fn = "imageCover%d.png" % i
            media.append(("xl/media/" + fn, png))
            rels.append('<Relationship Id="%s" Type="%s/image" Target="../media/%s"/>' % (rid, NS_R, fn))
            anchors.append(
                '<xdr:oneCellAnchor>'
                '<xdr:from><xdr:col>%d</xdr:col><xdr:colOff>120000</xdr:colOff>'
                '<xdr:row>%d</xdr:row><xdr:rowOff>70000</xdr:rowOff></xdr:from>'
                '<xdr:ext cx="%d" cy="%d"/>'
                '<xdr:pic><xdr:nvPicPr>'
                '<xdr:cNvPr id="%d" name="sign%d"/>'
                '<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
                '<xdr:blipFill><a:blip xmlns:r="%s" r:embed="%s"/>'
                '<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
                '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="%d" cy="%d"/></a:xfrm>'
                '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:pic>'
                '<xdr:clientData/></xdr:oneCellAnchor>'
                % (ci, ri, cx, cy, i + 1, i, NS_R, rid, cx, cy)
            )
        drawing_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<xdr:wsDr xmlns:xdr="%s" xmlns:a="%s">%s</xdr:wsDr>'
            % (NS_XDR, NS_A, "".join(anchors))
        )
        drawing_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">%s</Relationships>'
            % "".join(rels)
        )
        return drawing_xml, drawing_rels, media

    def __build_filled(self, src_path, values, imgs=None):
        """在 zip 层复制整个 xlsx：改「封面」页文字单元格；如有签名图则给封面挂一个 drawing 插入签名章。
        其余内容（bug 描述截图等）原样保留。纯文件操作，可放线程池执行。"""
        imgs = imgs or []
        DRAW_PART = "xl/drawings/drawingCover.xml"
        DRAW_RELS = "xl/drawings/_rels/drawingCover.xml.rels"
        with zipfile.ZipFile(src_path, "r") as zin:
            wb = zin.read("xl/workbook.xml").decode("utf-8", errors="ignore")
            m = re.search(r'<sheet name="封面"[^>]*r:id="([^"]+)"', wb)
            rid = m.group(1) if m else "rId1"
            wrels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="ignore")
            mt = re.search(r'<Relationship Id="' + re.escape(rid) + r'"[^>]*Target="([^"]+)"', wrels)
            target = mt.group(1) if mt else "worksheets/sheet1.xml"
            sheet_path = target.lstrip("/") if target.startswith("/") else "xl/" + target
            sheet_name = sheet_path.rsplit("/", 1)[-1]
            sheet_rels_path = "xl/worksheets/_rels/" + sheet_name + ".rels"

            names = set(zin.namelist())
            drawing_xml = drawing_rels = ""
            media = []
            draw_rid = None
            if imgs:
                drawing_xml, drawing_rels, media = self.__drawing_xml(imgs)
                existing = re.findall(r'Id="rId(\d+)"', zin.read(sheet_rels_path).decode("utf-8", errors="ignore")) if sheet_rels_path in names else []
                draw_rid = "rId%d" % (max([int(x) for x in existing], default=0) + 1)

            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == sheet_path:
                        xml = data.decode("utf-8", errors="ignore")
                        for coord, val in values.items():
                            xml = self.__set_cell(xml, coord, val)
                        if imgs:
                            xml = xml.replace("</worksheet>", '<drawing r:id="%s"/></worksheet>' % draw_rid)
                        data = xml.encode("utf-8")
                    elif item.filename == sheet_rels_path and imgs:
                        rels_xml = data.decode("utf-8", errors="ignore")
                        rel = ('<Relationship Id="%s" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawingCover.xml"/>' % draw_rid)
                        data = rels_xml.replace("</Relationships>", rel + "</Relationships>").encode("utf-8")
                    elif item.filename == "[Content_Types].xml" and imgs:
                        ct = data.decode("utf-8", errors="ignore")
                        ov = '<Override PartName="/xl/drawings/drawingCover.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                        data = ct.replace("</Types>", ov + "</Types>").encode("utf-8")
                    zout.writestr(item, data)

                if imgs:
                    # 若封面原本没有 rels 文件，需新建（含 drawing 关系）
                    if sheet_rels_path not in names:
                        rel = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                               '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                               '<Relationship Id="%s" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawingCover.xml"/>'
                               '</Relationships>' % (draw_rid or "rId1"))
                        zout.writestr(sheet_rels_path, rel)
                    zout.writestr(DRAW_PART, drawing_xml)
                    zout.writestr(DRAW_RELS, drawing_rels)
                    for mname, mbytes in media:
                        zout.writestr(mname, mbytes)
        return buf.getvalue()

    def __dhf_file_no(self, prod_id):
        if not prod_id:
            return ""
        row = db.session.execute(
            select(ProdDhf).where(ProdDhf.prod_id == prod_id, ProdDhf.name == DOC_NAME).order_by(ProdDhf.id.asc())
        ).scalars().first()
        if not row:
            row = db.session.execute(
                select(ProdDhf).where(ProdDhf.prod_id == prod_id, ProdDhf.name.like(f"%{DOC_NAME}%")).order_by(ProdDhf.id.asc())
            ).scalars().first()
        return (row.code or "").strip() if row and row.code else ""

    @staticmethod
    def __norm_level(v):
        t = re.sub(r"\s+", "", str(v or ""))
        for lv in LEVELS:
            if lv in t:
                return lv
        m = re.search(r"([1-5一二三四五])", t)
        if m:
            mp = {"1": "一级", "2": "二级", "3": "三级", "4": "四级", "5": "五级",
                  "一": "一级", "二": "二级", "三": "三级", "四": "四级", "五": "五级"}
            return mp.get(m.group(1), "")
        return ""

    def __parse_stats(self, file_bytes):
        """解析「bug描述」页：按级别统计 Bug总数/已解决/遗留。"""
        stats = {
            "total": 0,
            "resolved": 0,
            "remaining": 0,
            "by_level": {lv: 0 for lv in LEVELS},
            "resolved_by_level": {lv: 0 for lv in LEVELS},
            "remaining_by_level": {lv: 0 for lv in LEVELS},
        }
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception:
            logger.exception("bug xlsx load failed")
            return stats
        ws = None
        for name in wb.sheetnames:
            if "bug" in name.lower() and ("描述" in name or "list" in name.lower()):
                ws = wb[name]
                break
        if ws is None:
            for name in wb.sheetnames:
                if "描述" in name:
                    ws = wb[name]
                    break
        if ws is None:
            return stats
        rows = list(ws.iter_rows(values_only=True))
        # 找表头行（含 bug编号/级数/状态）
        hdr_idx, cols = -1, {}
        for i, r in enumerate(rows[:6]):
            cells = [re.sub(r"\s+", "", str(c or "")) for c in r]
            joined = " ".join(cells)
            if ("bug编号" in joined.lower() or "bug编号" in joined) and ("级" in joined):
                hdr_idx = i
                for j, c in enumerate(cells):
                    cl = c.lower()
                    if "bug编号" in cl or "bug编号" in c:
                        cols["code"] = j
                    elif "级数" in c or "级别" in c or c == "级":
                        cols["level"] = j
                    elif "状态" in c:
                        cols["status"] = j
                break
        if hdr_idx < 0 or "level" not in cols:
            return stats
        for r in rows[hdr_idx + 1:]:
            code = str(r[cols["code"]] if cols.get("code") is not None and cols["code"] < len(r) else "").strip()
            if not code:
                continue
            lv = self.__norm_level(r[cols["level"]] if cols["level"] < len(r) else "")
            if not lv:
                continue
            status = str(r[cols["status"]] if cols.get("status") is not None and cols["status"] < len(r) else "")
            resolved = any(k in status for k in RESOLVED_KW)
            stats["total"] += 1
            stats["by_level"][lv] += 1
            if resolved:
                stats["resolved"] += 1
                stats["resolved_by_level"][lv] += 1
            else:
                stats["remaining"] += 1
                stats["remaining_by_level"][lv] += 1
        return stats

    def __to_obj(self, row: BugDoc, product: Product = None):
        obj = BugDocObj(id=row.id, product_id=row.product_id, version=row.version,
                        file_no=row.file_no, change_log=row.change_log,
                        file_name=row.file_name, stats=row.stats, create_time=row.create_time)
        if not (obj.file_no or "").strip():
            obj.file_no = self.__dhf_file_no(row.product_id)
        if product:
            obj.product_name = product.name
            obj.product_version = product.full_version
            obj.product_full_version = product.full_version
            obj.product_type_code = product.type_code
        return obj

    async def add_bug_doc(self, form: BugDocForm, file=None):
        try:
            sql = select(func.count(BugDoc.id)).where(BugDoc.product_id == form.product_id, BugDoc.version == form.version)
            if db.session.execute(sql).scalar() > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            row = BugDoc(product_id=form.product_id, version=form.version, file_no=form.file_no, change_log=form.change_log)
            if file is not None:
                data = await file.read()
                row.file_name = file.filename
                row.file_path = self.__save_to_disk(data, file.filename)
                row.file_data = None
                row.stats = self.__parse_stats(data)
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok(data=BugDocForm(id=row.id))
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def update_bug_doc(self, form: BugDocForm, file=None):
        try:
            row: BugDoc = db.session.execute(select(BugDoc).where(BugDoc.id == form.id)).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key in ("product_id", "version", "file_no", "change_log"):
                val = getattr(form, key, None)
                if val is not None:
                    setattr(row, key, val)
            if file is not None:
                data = await file.read()
                row.file_name = file.filename
                row.file_path = self.__save_to_disk(data, file.filename)
                row.file_data = None
                row.stats = self.__parse_stats(data)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def delete_bug_doc(self, id: int):
        db.session.execute(delete(BugDoc).where(BugDoc.id == id))
        db.session.commit()
        return Resp.resp_ok()

    async def get_bug_doc(self, id: int):
        sql = select(BugDoc, Product).join(Product, BugDoc.product_id == Product.id).where(BugDoc.id == id)
        row = db.session.execute(sql).first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        doc, product = row
        return Resp.resp_ok(data=self.__to_obj(doc, product))

    async def list_bug_doc(self, op_user: UserObj = None, product_id: int = 0, version: str = None, page_index: int = 0, page_size: int = 10):
        wheres = []
        if product_id:
            wheres.append(BugDoc.product_id == product_id)
        if version:
            wheres.append(BugDoc.version.like(f"%{version}%"))
        if op_user and op_user.id != 1 and op_user.role_code == Roles.product_manager.value.code:
            wheres.append(Product.create_user_id == op_user.id)
        sql_total = select(func.count(BugDoc.id)).join(Product, BugDoc.product_id == Product.id).where(*wheres)
        total = db.session.execute(sql_total).scalar() or 0
        # 列表不加载 file_data，避免大文件拖慢
        sql = (select(BugDoc.id, BugDoc.product_id, BugDoc.version, BugDoc.file_no, BugDoc.change_log,
                      BugDoc.file_name, BugDoc.stats, BugDoc.create_time, Product)
               .join(Product, BugDoc.product_id == Product.id).where(*wheres)
               .order_by(BugDoc.id.desc()).offset(page_index * page_size).limit(page_size))
        rows: List[BugDocObj] = []
        for r in db.session.execute(sql).all():
            product = r[-1]
            obj = BugDocObj(id=r[0], product_id=r[1], version=r[2], file_no=r[3], change_log=r[4],
                            file_name=r[5], stats=r[6], create_time=r[7])
            if not (obj.file_no or "").strip():
                obj.file_no = self.__dhf_file_no(r[1])
            if product:
                obj.product_name = product.name
                obj.product_version = product.full_version
                obj.product_full_version = product.full_version
                obj.product_type_code = product.type_code
            rows.append(obj)
        return Resp.resp_ok(data=Page(total=total, rows=rows, page_index=page_index, page_size=page_size))

    async def download_bug_doc(self, id: int):
        """下载：在原始 xlsx 基础上自动回填封面（编制/审核/批准人、日期、生效日期、文件修订记录），
        返回已回填文件的磁盘路径（带缓存）。回填只改「封面」页，bug 描述里的截图等内容原样保留；出错回退原文件。"""
        row: BugDoc = db.session.execute(select(BugDoc).where(BugDoc.id == id)).scalars().first()
        if not row:
            return None, None
        path = self.__ensure_disk(row)
        if not path:
            return None, None
        filename = row.file_name or f"{DOC_NAME}.xlsx"
        try:
            values, imgs, _ = self.__cover_plan(row)  # 纯 DB 读取
            filled = await run_in_threadpool(self.__ensure_filled, path, values, imgs)  # 生成/取缓存，放线程池
            return filename, filled
        except Exception:
            logger.exception("fill bug cover failed, fallback to original")
            return filename, path

    @staticmethod
    def __coord_rc(coord):
        """A1 式坐标 → (行索引, 列索引)，均 0 起。"""
        col = "".join(ch for ch in coord if ch.isalpha())
        rn = int("".join(ch for ch in coord if ch.isdigit()))
        ci = 0
        for ch in col:
            ci = ci * 26 + (ord(ch.upper()) - 64)
        return rn - 1, ci - 1

    async def preview_bug_doc(self, id: int):
        """在线预览：把上传的 xlsx 各页签解析成文字表格返回（不含图片）。
        「封面」页会叠加自动回填的编制/审核/批准人、日期、生效日期、文件修订记录，与下载内容一致。"""
        row: BugDoc = db.session.execute(select(BugDoc).where(BugDoc.id == id)).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        path = self.__ensure_disk(row)
        if not path:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        try:
            cover_vals, _imgs, sign_urls = self.__cover_plan(row)
            cover_vals = dict(cover_vals)
            cover_vals.update(sign_urls)  # 有签名图的单元格直接放 data URL，前端渲染成签名章
        except Exception:
            logger.exception("cover values for preview failed")
            cover_vals = {}
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            logger.exception("bug xlsx preview failed")
            return Resp.resp_err(msg="文件解析失败")
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows, maxc = [], 0
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i >= 2000:
                    break
                cells = ["" if c is None else str(c) for c in r]
                while cells and cells[-1] == "":
                    cells.pop()
                rows.append(cells)
                maxc = max(maxc, len(cells))
            # 去掉全空尾行
            while rows and not any(x.strip() for x in rows[-1]):
                rows.pop()
            for rr in rows:
                while len(rr) < maxc:
                    rr.append("")
            # 「封面」页叠加自动回填值（仅填空白单元格）
            if name == "封面" and cover_vals:
                for coord, val in cover_vals.items():
                    ri, ci = self.__coord_rc(coord)
                    while len(rows) <= ri:
                        rows.append([""] * maxc)
                    while len(rows[ri]) <= ci:
                        rows[ri].append("")
                    if not str(rows[ri][ci]).strip():
                        rows[ri][ci] = val
            sheets.append({"name": name, "rows": rows})
        return Resp.resp_ok(data={"sheets": sheets, "file_name": row.file_name or ""})

    def template_bytes(self):
        try:
            with open(TEMPLATE_PATH, "rb") as f:
                return f.read()
        except Exception:
            logger.exception("bug template read failed")
            return None

    def stats_for_product(self, prod_id):
        """取该产品最新一份 Bug管理文档的缺陷统计（供软件测试报告引用）。"""
        if not prod_id:
            return None
        row = db.session.execute(
            select(BugDoc.stats).where(BugDoc.product_id == prod_id).order_by(BugDoc.id.desc())
        ).scalars().first()
        return row or None
