import logging
import json
import os
import re
from typing import List
from sqlalchemy import select, delete, func
from sqlalchemy.sql import desc
from sqlalchemy.dialects.postgresql import insert as pg_insert
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from ..obj.vobj_user import UserObj
from ..model.test_set import TestSet
from ..model.srs_doc import SrsDoc, SrsNode
from ..model.srs_req import ReqRcm
from ..model.product import Product, UserProd
from ..model.test_case import TestCase
from ..model.rcm import Rcm
from ..model.srs_req import SrsReq
from ..model.prod_rcm import ProdRcm
from ..obj.tobj_prod_rcm import ProdRcmsForm
from ..obj.vobj_prod_rcm import ProdRcmObj
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):

    NO_TEST_PROOF_MAP = {
        "RCM019": "模型文件",
        "RCM056": "软件配置管理计划",
        "RCM057": "工具确认计划/工具确认报告",
        "RCM058": "生产过程控制程序",
        "RCM059": "包装运输检测报告",
        "RCM061": "成品检验作业指导书",
        "RCM062": "库房管理制度",
        "RCM097": "培训记录",
        "RCM100": "培训记录",
        "RCM101": "安装确认报告",
        "RCM126": "网络安全扫描报告",
    }

    def __normalize_code(self, value: str) -> str:
        text = re.sub(r"\s+", "", str(value or "")).replace("＿", "_").upper()
        return re.sub(r"[，。；;、,.]+$", "", text)

    def __split_srs_codes(self, value: str) -> List[str]:
        text = str(value or "")
        matches = re.findall(r"SRS[\s\-_]*[A-Z0-9.]+(?:\s*-\s*[A-Z0-9.]+)*", text, flags=re.I)
        if matches:
            codes = [self.__normalize_code(match.strip(" ,，;；、\n\r\t")) for match in matches]
        else:
            codes = [self.__normalize_code(item) for item in re.split(r"[,，;；、\s]+", text)]
        return [code for code in dict.fromkeys(codes) if code]

    def __split_rcm_codes(self, value: str) -> List[str]:
        text = str(value or "")
        matches = re.findall(r"\bRCM[\s\-_]*[A-Z0-9]+(?:[\-_][A-Z0-9]+)*\b", text, flags=re.I)
        codes = [self.__normalize_code(match.strip(" ,，;；、\n\r\t")) for match in matches]
        return [code for code in dict.fromkeys(codes) if code]

    def __build_trace_rule_from_srs_code(self, srs_code: str):
        code = str(srs_code or "").strip().upper()
        matched = re.match(r"^SRS-([A-Z]+)(\d+)-(\d+)$", code)
        if not matched:
            return None
        major = matched.group(2)
        minor = matched.group(3)
        if len(major) < 2:
            return None
        return {
            "if_code": major[-2:],
            "unit_group": minor.zfill(3),
        }

    def __normalize_test_stage(self, stage: str) -> str:
        text = str(stage or "").strip()
        if "系统" in text:
            return "系统测试"
        if "集成" in text:
            return "集成测试"
        if "单元" in text:
            return "单元测试"
        if "用户" in text:
            return "用户测试"
        return text

    def __normalize_match_text(self, value: str) -> str:
        text = re.sub(r"\bRCM[\s\-_]*[A-Z0-9]+(?:[\-_][A-Z0-9]+)*[.。]?", "", str(value or ""), flags=re.I)
        return re.sub(r"[\s\u3000，。；;、,.：:（）()【】\[\]_/\\!！\"“”‘’\-]+", "", text).upper()

    def __rcm_match_phrases(self, description: str) -> List[str]:
        text = re.sub(r"\bRCM[\s\-_]*[A-Z0-9]+(?:[\-_][A-Z0-9]+)*[.。]?", "", str(description or ""), flags=re.I)
        raw_parts = [item for item in re.split(r"[，。；;、,.：:（）()【】\[\]\s]+", text) if item]
        phrases = []
        stop_words = [
            "软件",
            "系统",
            "产品",
            "本产品",
            "前端页面",
            "页面",
            "增加",
            "新增",
            "提示",
            "校验",
            "验证",
        ]
        for part in raw_parts:
            normalized = self.__normalize_match_text(part)
            if len(normalized) >= 6:
                phrases.append(normalized)
            reduced = normalized
            for word in stop_words:
                reduced = reduced.replace(self.__normalize_match_text(word), "")
            if len(reduced) >= 6:
                phrases.append(reduced)
            for size in range(min(len(reduced), 14), 5, -1):
                for idx in range(0, len(reduced) - size + 1):
                    phrases.append(reduced[idx:idx + size])
        normalized_text = self.__normalize_match_text(text)
        if "界面" in normalized_text and ("中文" in normalized_text or "英文" in normalized_text):
            phrases.append("界面语言")
        if "DICOM标签" in normalized_text:
            if "不完整" in normalized_text:
                phrases.extend(["DICOM标签不完整", "不完整数据"])
            if "错误" in normalized_text:
                phrases.extend(["DICOM标签错误", "错误数据"])
        phrases = [phrase for phrase in dict.fromkeys(phrases) if len(phrase) >= 4]
        phrases.sort(key=len, reverse=True)
        return phrases

    def __rcm_focus_match_sets(self, description: str) -> List[List[str]]:
        normalized = self.__normalize_match_text(description)
        focus_sets = []

        unit_terms = []
        for unit in re.findall(r"单位为([A-Z0-9一-龥]+)", normalized, flags=re.I):
            if unit:
                unit_terms.append(f"单位为{unit}")

        for subject in ["长度测量", "角度测量", "CT点值"]:
            if subject in normalized:
                focus_sets.append([subject, *unit_terms] if unit_terms else [subject])

        for feature in re.findall(r"(?:提供|支持|具备|包含)(.+?功能)", normalized):
            feature = re.sub(r"^(软件|系统|产品|增加|新增)+", "", feature)
            feature_term = feature if feature.endswith("功能") else f"{feature}功能"
            if len(feature) >= 2:
                feature_core = feature_term[:-2] if feature_term.endswith("功能") else feature_term
                if len(feature_core) >= 2:
                    focus_sets.append([feature_core])
                focus_sets.append(["提供", feature_term])

        if "DICOM标签" in normalized:
            alternatives = []
            if "不完整" in normalized:
                alternatives.append("不完整")
            if "错误" in normalized:
                alternatives.append("错误")
            if alternatives:
                focus_sets.extend([["DICOM标签", alternative] for alternative in alternatives])
        if "DICOM" in normalized and ("标准" in normalized or "DICOM30" in normalized):
            focus_terms = ["DICOM"]
            if "标准" in normalized:
                focus_terms.append("标准")
            if "筛除" in normalized or "过滤" in normalized or "不满足" in normalized or "不符合" in normalized:
                focus_terms.append("过滤")
            focus_sets.append(focus_terms)
        if "磁盘" in normalized and ("容量" in normalized or "空间" in normalized) and "不足" in normalized:
            focus_sets.append(["磁盘", "空间不足"])
        if "备份" in normalized:
            if "数据库" in normalized:
                focus_sets.append(["数据库", "备份"])
            if "磁盘" in normalized or "图像数据" in normalized:
                focus_sets.append(["DCM", "备份"])
        if ("传输异常" in normalized or "断电" in normalized or "宕机" in normalized) and "图像数据" in normalized and "丢失" in normalized:
            if "传输异常" in normalized:
                focus_sets.append(["传输异常", "图像数据", "丢失"])
            if "断电" in normalized or "宕机" in normalized:
                focus_sets.append(["断电", "图像数据", "丢失"])
        if "培训" in normalized:
            focus_sets.append(["培训"])
        if "版本" in normalized and ("确认" in normalized or "校验" in normalized or "检查" in normalized):
            focus_sets.append(["版本"])
        if "安全扫描" in normalized or "补丁" in normalized or "漏洞" in normalized:
            if "安全扫描" in normalized:
                focus_sets.append(["安全扫描"])
            if "补丁" in normalized:
                focus_sets.append(["补丁"])
            if "漏洞" in normalized:
                focus_sets.append(["漏洞"])
        if "输入框" in normalized and ("格式" in normalized or "要求" in normalized or "定义" in normalized):
            if "账号" in normalized:
                focus_sets.extend([
                    ["账号输入框"],
                    ["请输入正确的账号"],
                    ["请输入账号"],
                    ["账号长度"],
                    ["账号错误字符"],
                    ["账号大小写"],
                    ["账号不能为空"],
                ])
            if "密码" in normalized:
                focus_sets.extend([
                    ["密码输入框"],
                    ["新密码输入框"],
                    ["确认新密码输入框"],
                    ["密码长度必须", "输入框"],
                    ["密码长度必须", "显示提示"],
                    ["密码格式", "输入框"],
                    ["密码格式", "显示提示"],
                ])

        if "用户分层" in normalized:
            focus_sets.append(["用户分层"])
        if "重启" in normalized and ("容错" in normalized or "异常" in normalized):
            focus_sets.append(["重启机器"])
        if "版权" in normalized and ("序列码" in normalized or "序列号" in normalized):
            focus_sets.append(["版权保护"])
        if "明文" in normalized:
            focus_sets.append(["明文"])
        if "目标物" in normalized and "编辑" in normalized:
            focus_sets.append(["编辑目标物"])

        return [terms for terms in focus_sets if terms]

    def __rcm_fuzzy_terms(self, description: str) -> List[str]:
        text = self.__normalize_match_text(description)
        stop_words = [
            "软件",
            "系统",
            "产品",
            "本产品",
            "用户",
            "提供",
            "支持",
            "具备",
            "包含",
            "增加",
            "新增",
            "功能",
            "提示",
            "信息",
            "校验",
            "验证",
            "进行",
            "要求",
            "设计",
            "操作",
            "页面",
        ]
        for word in stop_words:
            text = text.replace(self.__normalize_match_text(word), " ")
        terms = []
        for part in re.split(r"\s+", text):
            if len(part) >= 2:
                terms.append(part)
            if len(part) > 6:
                for size in range(min(len(part), 6), 1, -1):
                    for idx in range(0, len(part) - size + 1):
                        terms.append(part[idx:idx + size])
        terms = [term for term in dict.fromkeys(terms) if len(term) >= 2]
        terms.sort(key=len, reverse=True)
        return terms

    def __filter_tests_by_rcm_fuzzy(self, tests: List[TestCase], rcm: Rcm = None) -> List[TestCase]:
        if not tests or not rcm:
            return []
        terms = self.__rcm_fuzzy_terms(getattr(rcm, "description", "") or "")
        if not terms:
            return []
        scored = []
        for test in tests:
            merged_text = "\n".join([
                str(getattr(test, "function", "") or ""),
                str(getattr(test, "description", "") or ""),
                str(getattr(test, "precondition", "") or ""),
                str(getattr(test, "test_step", "") or ""),
                str(getattr(test, "expect", "") or ""),
                str(getattr(test, "note", "") or ""),
            ])
            haystack = self.__normalize_match_text(merged_text)
            matched_terms = [term for term in terms if term in haystack]
            if matched_terms:
                scored.append((sum(len(term) for term in matched_terms), len(matched_terms), test))
        if not scored:
            return []
        max_score = max(score for score, _, _ in scored)
        max_count = max(count for score, count, _ in scored if score == max_score)
        matched_tests = [test for score, count, test in scored if score == max_score and count == max_count]
        if len(matched_tests) > 6:
            return []
        return matched_tests

    def __filter_tests_by_rcm_description(self, tests: List[TestCase], rcm: Rcm = None, require_match: bool = False) -> List[TestCase]:
        if not tests or not rcm:
            return [] if require_match else tests
        rcm_description = getattr(rcm, "description", "") or ""
        phrases = self.__rcm_match_phrases(rcm_description)
        if not phrases:
            return [] if require_match else tests

        focused_match_sets = self.__rcm_focus_match_sets(rcm_description)

        scored = []
        focused_matched_tests = []
        for test in tests:
            merged_text = "\n".join([
                str(getattr(test, "function", "") or ""),
                str(getattr(test, "description", "") or ""),
                str(getattr(test, "precondition", "") or ""),
                str(getattr(test, "test_step", "") or ""),
                str(getattr(test, "expect", "") or ""),
                str(getattr(test, "note", "") or ""),
            ])
            haystack = self.__normalize_match_text(merged_text)
            if focused_match_sets and any(all(term in haystack for term in match_set) for match_set in focused_match_sets):
                focused_matched_tests.append(test)
                continue
            matched_phrases = [phrase for phrase in phrases if phrase in haystack]
            if matched_phrases:
                scored.append((sum(len(phrase) for phrase in matched_phrases), len(matched_phrases), max(len(phrase) for phrase in matched_phrases), test))
        if focused_match_sets:
            return focused_matched_tests
        if not scored:
            return [] if require_match else tests

        max_phrase_len = max(phrase_len for _, _, phrase_len, _ in scored)
        matched_tests = [test for _, _, phrase_len, test in scored if phrase_len == max_phrase_len]
        if len(matched_tests) > 6:
            return [] if require_match else tests
        return matched_tests

    def __format_test_code_ranges(self, test_codes: List[str]) -> List[str]:
        groups = dict()
        order = []
        for code in test_codes:
            text = str(code or "").strip()
            matched = re.match(r"^(.*?)(\d+)$", text)
            if not matched:
                order.append(("raw", text))
                continue
            prefix = matched.group(1)
            if prefix not in groups:
                groups[prefix] = []
                order.append(("group", prefix))
            groups[prefix].append(text)

        result = []
        for item_type, value in order:
            if item_type == "raw":
                result.append(value)
                continue
            group = groups.get(value) or []
            if len(group) == 1:
                result.append(group[0])
            elif group:
                result.append(f"{group[0]}~{group[-1]}")
        return result

    async def add_prod_rcms(self, form: ProdRcmsForm):
        try:
            values = []
            for rcm_id in form.rcm_ids:
                values.append(dict(prod_id=form.prod_id, rcm_id=rcm_id))
            db.session.execute(pg_insert(ProdRcm).values(values).on_conflict_do_nothing())
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_prod_rcms(self, ids: List[str]):
        if ids:
            db.session.execute(delete(ProdRcm).where(ProdRcm.id.in_(ids)))
            db.session.commit()
        return Resp.resp_ok()

    def __repair_req_rcms_from_srs_nodes(self, product_ids: List[int]):
        product_ids = [product_id for product_id in dict.fromkeys(product_ids or []) if product_id]
        if not product_ids:
            return
        rows = db.session.execute(
            select(SrsNode, SrsDoc)
            .join(SrsDoc, SrsNode.doc_id == SrsDoc.id)
            .where(SrsDoc.product_id.in_(product_ids))
        ).all()
        if not rows:
            return

        doc_ids = [doc.id for _node, doc in rows]
        all_nodes = db.session.execute(select(SrsNode).where(SrsNode.doc_id.in_(doc_ids))).scalars().all()
        node_by_doc = {}
        for node in all_nodes:
            node_by_doc.setdefault(node.doc_id, {})[node.n_id] = node

        def normalize_match_text(value: str):
            txt = re.sub(r"^\s*\d+(?:\.\d+)*[\s、.．:：\-]*", "", str(value or "")).strip()
            return re.sub(r"[\s\u3000、，。；;：:（）()【】\[\]_\-]+", "", txt).upper()

        req_rows_all = db.session.execute(select(SrsReq).where(SrsReq.doc_id.in_(doc_ids))).scalars().all()
        req_name_index = {}
        for req in req_rows_all:
            for name in [req.sub_function or "", req.function or "", req.module or ""]:
                key = normalize_match_text(name)
                if not key:
                    continue
                codes = req_name_index.setdefault((req.doc_id, key), [])
                if req.code not in codes:
                    codes.append(req.code)

        def resolve_srs_by_function_context(node):
            cur = node
            safety = 0
            titles = []
            while cur and safety < 50:
                if cur.title:
                    titles.append(cur.title)
                cur = node_by_doc.get(node.doc_id, {}).get(cur.p_id)
                safety += 1
            for title in titles:
                key = normalize_match_text(title)
                codes = req_name_index.get((node.doc_id, key)) or []
                if len(codes) == 1:
                    return self.__normalize_code(codes[0])
            return ""

        def node_search_text(node):
            parts = [node.title or "", node.label or "", node.text or ""]
            if node.table:
                parts.append(table_search_text(node.table))
            return "\n".join(parts)

        def table_search_text(table):
            payload = table
            if isinstance(payload, str):
                try:
                    payload = json.loads(payload)
                except Exception:
                    return payload
            if not isinstance(payload, dict):
                return str(payload or "")
            values = []
            for header in payload.get("headers") or []:
                values.append(str((header or {}).get("name") or ""))
            for row in payload.get("rows") or []:
                for value in (row or {}).values():
                    values.append(str(value or ""))
            for row in payload.get("cells") or []:
                for cell in row or []:
                    values.append(str((cell or {}).get("value") or ""))
            return "\n".join(values)

        def extract_srs_codes_from_node(node):
            result = []
            for hit in re.findall(r"SRS[\s\-_]*[A-Z0-9.]+(?:\s*-\s*[A-Z0-9.]+)*", node_search_text(node), flags=re.I):
                code = self.__normalize_code(hit)
                if code and code not in result:
                    result.append(code)
            return result

        pairs = []
        all_srs_codes = set()
        all_rcm_codes = set()
        for node, _doc in rows:
            direct_srs_codes = extract_srs_codes_from_node(node)
            srs_code = self.__normalize_code(node.srs_code)
            parent_id = node.p_id
            if not direct_srs_codes:
                while not srs_code and parent_id:
                    parent = node_by_doc.get(node.doc_id, {}).get(parent_id)
                    if not parent:
                        break
                    srs_code = self.__normalize_code(parent.srs_code)
                    parent_id = parent.p_id
            if not srs_code:
                direct_srs_codes = direct_srs_codes or []
            else:
                direct_srs_codes = direct_srs_codes or [srs_code]
            rcm_codes = self.__split_rcm_codes(node.rcm_codes) or self.__split_rcm_codes(node_search_text(node))
            if not rcm_codes:
                continue
            if not direct_srs_codes:
                context_srs_code = resolve_srs_by_function_context(node)
                direct_srs_codes = [context_srs_code] if context_srs_code else []
            for code in direct_srs_codes:
                if not code:
                    continue
                all_srs_codes.add(code)
                all_rcm_codes.update(rcm_codes)
                pairs.append((node.doc_id, code, rcm_codes))
        if not pairs:
            return

        req_rows = db.session.execute(select(SrsReq).where(SrsReq.code.in_(all_srs_codes))).scalars().all()
        req_id_by_doc_code = {(row.doc_id, self.__normalize_code(row.code)): row.id for row in req_rows}
        rcm_rows = db.session.execute(select(Rcm)).scalars().all()
        rcm_id_by_code = {self.__normalize_code(row.code): row.id for row in rcm_rows}
        affected_rcm_ids = [rcm_id_by_code[code] for code in all_rcm_codes if code in rcm_id_by_code]

        insert_values = []
        seen_pairs = set()
        for doc_id, srs_code, rcm_codes in pairs:
            req_id = req_id_by_doc_code.get((doc_id, srs_code))
            if not req_id:
                continue
            for rcm_code in rcm_codes:
                rcm_id = rcm_id_by_code.get(rcm_code)
                if not rcm_id:
                    continue
                key = (req_id, rcm_id)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)
                insert_values.append(dict(req_id=req_id, rcm_id=rcm_id))
        if affected_rcm_ids:
            affected_req_ids = db.session.execute(
                select(SrsReq.id).where(SrsReq.doc_id.in_(doc_ids))
            ).scalars().all()
            if affected_req_ids:
                db.session.execute(
                    delete(ReqRcm).where(
                        ReqRcm.req_id.in_(affected_req_ids),
                        ReqRcm.rcm_id.in_(affected_rcm_ids),
                    )
                )
        if insert_values:
            db.session.execute(pg_insert(ReqRcm).values(insert_values).on_conflict_do_nothing())
        if affected_rcm_ids or insert_values:
            db.session.commit()
    
    def __query_srs_reqs(self, rcm_ids: List[int]) -> List[str]:
        rcm_ids = [rcm_id for rcm_id in (rcm_ids or []) if rcm_id]
        if not rcm_ids:
            return [], {}
        sql = select(ReqRcm, SrsReq, SrsDoc).join(SrsReq, ReqRcm.req_id == SrsReq.id)
        sql = sql.join(SrsDoc, SrsReq.doc_id == SrsDoc.id)
        sql = sql.where(ReqRcm.rcm_id.in_(rcm_ids)).order_by(SrsReq.code)
        results = dict()
        srs_codes = []        
        for req_rcm, row_srs, row_doc in db.session.execute(sql):
            key = (row_doc.product_id, req_rcm.rcm_id)
            reqs = results.get(key) or []
            reqs.append(row_srs)
            results[key] = reqs
            srs_codes.append(self.__normalize_code(row_srs.code))

        return srs_codes, results
    
    def __query_tests(self, product_ids: List[int], srs_codes: List[str]) -> dict:
        product_ids = [product_id for product_id in dict.fromkeys(product_ids or []) if product_id]
        wanted_srs_codes = {self.__normalize_code(code) for code in (srs_codes or []) if self.__normalize_code(code)}
        if not product_ids or not wanted_srs_codes:
            return {}
        sql = select(TestCase, TestSet).join(TestSet, TestCase.set_id == TestSet.id)
        sql = sql.where(TestSet.product_id.in_(product_ids)).order_by(TestCase.set_id, TestCase.code)
        rows: list[TestCase, TestSet] = db.session.execute(sql).all()
        results = dict()
        for row_test, row_set in rows:
            setattr(row_test, "_prod_rcm_stage", self.__normalize_test_stage(row_set.stage))
            for srs_code in self.__split_srs_codes(row_test.srs_code):
                if srs_code in wanted_srs_codes:
                    results.setdefault((row_set.product_id, srs_code), []).append(row_test)
        for product_id in product_ids:
            for srs_code in wanted_srs_codes:
                trace_rule = self.__build_trace_rule_from_srs_code(srs_code)
                if not trace_rule:
                    continue
                result_key = (product_id, srs_code)
                current_tests = results.setdefault(result_key, [])
                current_codes = {self.__normalize_code(test.code) for test in current_tests}
                for stage, code_prefix in [
                    ("系统测试", "TS"),
                    ("集成测试", "TI"),
                    ("单元测试", "TU"),
                    ("用户测试", "TY"),
                ]:
                    if any(getattr(test, "_prod_rcm_stage", "") == stage for test in current_tests):
                        continue
                    prefix = f"{code_prefix}{trace_rule['if_code']}-{trace_rule['unit_group']}-"
                    matched_tests = [
                        row_test
                        for row_test, row_set in rows
                        if row_set.product_id == product_id
                        and getattr(row_test, "_prod_rcm_stage", "") == stage
                        and self.__normalize_code(row_test.code).startswith(prefix)
                        and self.__normalize_code(row_test.code) not in current_codes
                    ]
                    current_tests.extend(matched_tests)
                    current_codes.update(self.__normalize_code(test.code) for test in matched_tests)
        return results
    
    def __merge_tests(self, product_id, srs_codes: List[str], tests_dict: dict, rcm: Rcm = None) -> List[str]:
        test_sets = dict()
        uniq_sets = dict()
        for srs_code in srs_codes:
            tests = tests_dict.get((product_id, self.__normalize_code(srs_code))) or []
            for stage in ["系统测试", "集成测试", "单元测试", "用户测试"]:
                stage_tests = [test for test in tests if getattr(test, "_prod_rcm_stage", "") == stage]
                if not stage_tests:
                    continue
                if stage == "用户测试":
                    tests = stage_tests
                    break
                matched_tests = self.__filter_tests_by_rcm_description(stage_tests, rcm, require_match=True)
                if not matched_tests and not self.__rcm_focus_match_sets(getattr(rcm, "description", "") or ""):
                    matched_tests = self.__filter_tests_by_rcm_fuzzy(stage_tests, rcm)
                if matched_tests:
                    tests = matched_tests
                    break
            else:
                tests = []
            for test in tests:
                key = (srs_code, test.set_id)
                uniq_set = uniq_sets.setdefault(key, set())
                if test.code not in uniq_set:
                    uniq_set.add(test.code)
                    test_sets.setdefault(key, []).append(test.code)
        results = []
        for key, tests in test_sets.items():
            results.extend(self.__format_test_code_ranges(tests))
        return results

    async def list_prod_rcm(self, op_user: UserObj, export = False, prod_id: int = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 

        sql = select(ProdRcm, Product, Rcm).join(Product, ProdRcm.prod_id == Product.id).outerjoin(Rcm, ProdRcm.rcm_id == Rcm.id)
        if prod_id:
            sql = sql.where(ProdRcm.prod_id == prod_id)
        if not prod_id and op_user and op_user.id != 1:
            subquery = select(UserProd.product_id).where(UserProd.user_id == op_user.id).scalar_subquery()
            sql = sql.where(Product.id.in_(subquery))
        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.order_by(Rcm.code)
        rows: list[ProdRcm, Product, Rcm] = db.session.execute(sql).all()
        objs = []
        self.__repair_req_rcms_from_srs_nodes([row_prd.id for _, row_prd, _ in rows])
        all_srs_codes, reqs_dict = self.__query_srs_reqs([prod_rcm.rcm_id for prod_rcm, _, _ in rows])
        tests_dict = self.__query_tests([row_prd.id for _, row_prd, _ in rows], all_srs_codes)
        for row, row_prd, row_rcm in rows:
            obj = ProdRcmObj(**row_rcm.dict()) if row_rcm else ProdRcmObj()
            obj.id = row.id
            obj.rcm_id = row.rcm_id
            obj.create_time = row.create_time
            reqs = reqs_dict.get((row_prd.id, row.rcm_id)) or []
            obj.srs_codes = list(dict.fromkeys([self.__normalize_code(req.code) for req in reqs if self.__normalize_code(req.code)]))
            obj.srs_flag = True if obj.srs_codes else False
            fixed_proof = self.NO_TEST_PROOF_MAP.get(self.__normalize_code(obj.code))
            if fixed_proof:
                obj.test_codes = [fixed_proof]
            else:
                obj.test_codes = self.__merge_tests(row_prd.id, obj.srs_codes, tests_dict, row_rcm)
            
            obj.product_name = row_prd.name
            obj.product_version = row_prd.full_version
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))

    export_columns = [
        "code",
        "description",
        "srs_flag",
        "srs_codes",
        "test_codes",
        "proof",
        "note",
    ]

    async def export_prod_rcms(self, op_user: UserObj, output, *args, **kwargs):
        resp = await self.list_prod_rcm(op_user, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_prod_rcm.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            for cidx, key in enumerate(self.export_columns, 1):
                value = obj.get(key)
                if key == "srs_codes" or key == "test_codes":
                    value = "，\n".join(value)
                if key == "srs_flag":
                    value = ts("yes") if value else ts("no")
                ws.cell(row=ridx, column=cidx, value=value)

        align = Alignment(vertical='center', wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
        wb.save(output)
        output.seek(0)
        