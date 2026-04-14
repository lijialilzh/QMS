import logging
import io
from openpyxl import load_workbook
from ...obj.tobj_srs_doc import TabHeader, Table

logger = logging.getLogger(__name__)

async def read_excel(file, stream=False):
    tables = []
    try:
        if stream:
            bys = await file.read()
            wb = load_workbook(io.BytesIO(bys))
        else:
            wb = load_workbook(file)
        for sheet in wb.worksheets:
            headers = []
            rows = []
            for ridx, row in enumerate(sheet.iter_rows(max_col=sheet.max_column)):
                if ridx == 0:
                    for cidx, cell in enumerate(row):
                        if not cell.value:
                            break
                        headers.append(TabHeader(code=str(cidx), name=cell.value))
                    continue
                values = {str(cidx): cell.value for cidx, cell in enumerate(row) if cidx < len(headers)}
                rows.append(values)
            table = Table(name=sheet.title, headers=headers, rows=rows)
            tables.append(table)
    except Exception as e:
        logger.exception("")
    return tables
