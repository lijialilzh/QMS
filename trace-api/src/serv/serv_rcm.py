import logging
import io
import os
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from ..obj.vobj_rcm import RcmObj
from ..model.rcm import Rcm
from ..model.prod_rcm import ProdRcm
from ..obj.tobj_rcm import RcmForm
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)




class Server(object):
    @staticmethod
    def __to_str(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def __read_cell_value(cell, ws_raw):
        value = cell.value
        if value is not None and str(value).strip() != "":
            return value
        raw = ws_raw.cell(row=cell.row, column=cell.column).value
        if isinstance(raw, str) and raw.startswith("="):
            return raw
        return value

    async def add_rcm(self, form: RcmForm):
        try:
            sql = select(func.count(Rcm.id)).where(Rcm.code == form.code)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            
            row = Rcm(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def import_rcms(self, file):
        try:
            bys = await file.read()
            wb = load_workbook(io.BytesIO(bys), data_only=True)
            wb_raw = load_workbook(io.BytesIO(bys), data_only=False)
            ws = wb[wb.sheetnames[0]]
            ws_raw = wb_raw[wb_raw.sheetnames[0]]
            affected = 0
            for row in ws.iter_rows(min_row=2):
                vals = [self.__read_cell_value(cell, ws_raw) for cell in row[:len(self.export_columns)]]
                if not vals or all((v is None or str(v).strip() == "") for v in vals):
                    continue
                code = self.__to_str(vals[0])
                if not code:
                    continue
                data = dict(
                    code=code,
                    description=self.__to_str(vals[1]),
                    proof=self.__to_str(vals[2]),
                    note=self.__to_str(vals[3]),
                )
                existed = db.session.execute(select(Rcm).where(Rcm.code == code)).scalars().first()
                if existed:
                    for k, v in data.items():
                        setattr(existed, k, v)
                else:
                    db.session.add(Rcm(**data))
                affected += 1
            db.session.commit()
            return Resp.resp_ok(data={"count": affected})
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_rcm(self, id):
        db.session.execute(delete(Rcm).where(Rcm.id == id))
        db.session.execute(delete(ProdRcm).where(ProdRcm.rcm_id == id))
        db.session.commit()
        return Resp.resp_ok()
   
    async def update_rcm(self, form: RcmForm):
        try:
            sql = select(Rcm).where(Rcm.id == form.id)
            row:Rcm = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def get_rcm(self, id:str):
        sql = select(Rcm).where(Rcm.id == id)
        row = db.session.execute(sql).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        return Resp.resp_ok(data=RcmObj(**row.dict()))

    async def list_rcm(self, export = False, fuzzy: str = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 
    
        sql = select(Rcm)
        if fuzzy:
            sql = sql.where(
                or_(
                    Rcm.code.like(f"%{fuzzy}%"),
                    Rcm.description.like(f"%{fuzzy}%"),
                    Rcm.proof.like(f"%{fuzzy}%"),
                    Rcm.note.like(f"%{fuzzy}%"),
                )
            )
        
        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.offset(page_size * page_index).limit(page_size).order_by(Rcm.code)
        rows: list[Rcm] = db.session.execute(sql).scalars().all()
        objs = []
        for row in rows:
            obj = RcmObj(**row.dict())
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))

    export_columns = [
        "code",
        "description",
        "proof",
        "note",
    ]

    async def export_rcms(self, output, *args, **kwargs):
        resp = await self.list_rcm(export=True, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_rcm.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            for cidx, key in enumerate(self.export_columns, 1):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
                ws.cell(row=ridx, column=cidx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        wb.save(output)
        output.seek(0)
        