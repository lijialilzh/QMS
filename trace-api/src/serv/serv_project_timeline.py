import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy import select, delete, func
from sqlalchemy.sql import asc
from ..model.project_timeline import ProjectTimelineRow, ProjectTimelineCell
from ..obj.tobj_project_timeline import TimelineRowForm, TimelineCellForm
from ..obj.vobj_project_timeline import TimelineRowObj
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Resp
from . import msg_err_db

logger = logging.getLogger(__name__)

# 固定部门列（与「体系时间线」模板一致）
TIMELINE_DEPTS = [
    "产品部",
    "模型部",
    "数据部",
    "产品开发部-开发",
    "产品开发部-测试",
    "质量部",
    "三级通用技术文件",
]


class Server(object):

    async def list_timeline(self, prod_id: int):
        if not prod_id:
            return Resp.resp_ok(data={"depts": TIMELINE_DEPTS, "rows": []})
        sql = (
            select(ProjectTimelineRow)
            .where(ProjectTimelineRow.prod_id == prod_id)
            .order_by(asc(ProjectTimelineRow.sort_order), asc(ProjectTimelineRow.id))
        )
        rows: list[ProjectTimelineRow] = db.session.execute(sql).scalars().all()
        row_ids = [r.id for r in rows]
        cell_map: dict[int, dict] = {}
        if row_ids:
            csql = select(ProjectTimelineCell).where(ProjectTimelineCell.row_id.in_(row_ids))
            for c in db.session.execute(csql).scalars().all():
                cell_map.setdefault(c.row_id, {})[c.dept] = c.output_result or ""
        objs = []
        for r in rows:
            obj = TimelineRowObj(**r.dict(), cells=cell_map.get(r.id, {}))
            objs.append(obj)
        return Resp.resp_ok(data={"depts": TIMELINE_DEPTS, "rows": objs})

    async def add_timeline_row(self, form: TimelineRowForm):
        try:
            if not form.prod_id:
                return Resp.resp_err(msg=ts("msg_err_param"))
            if form.sort_order is None:
                max_sort = db.session.execute(
                    select(func.coalesce(func.max(ProjectTimelineRow.sort_order), 0)).where(
                        ProjectTimelineRow.prod_id == form.prod_id
                    )
                ).scalar()
                form.sort_order = int(max_sort or 0) + 1
            if not form.row_type:
                form.row_type = "date"
            row = ProjectTimelineRow(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok(data={"id": row.id})
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def update_timeline_row(self, form: TimelineRowForm):
        try:
            row: ProjectTimelineRow = db.session.execute(
                select(ProjectTimelineRow).where(ProjectTimelineRow.id == form.id)
            ).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key in ("id", "prod_id"):
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def delete_timeline_row(self, ids: list):
        ids = [int(i) for i in (ids or []) if str(i).strip().isdigit()]
        if not ids:
            return Resp.resp_err(msg=ts("msg_err_param"))
        try:
            db.session.execute(delete(ProjectTimelineCell).where(ProjectTimelineCell.row_id.in_(ids)))
            db.session.execute(delete(ProjectTimelineRow).where(ProjectTimelineRow.id.in_(ids)))
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def update_timeline_cell(self, form: TimelineCellForm):
        try:
            if not form.row_id or not form.dept:
                return Resp.resp_err(msg=ts("msg_err_param"))
            if form.dept not in TIMELINE_DEPTS:
                return Resp.resp_err(msg=ts("msg_err_param"))
            cell: ProjectTimelineCell = db.session.execute(
                select(ProjectTimelineCell).where(
                    ProjectTimelineCell.row_id == form.row_id, ProjectTimelineCell.dept == form.dept
                )
            ).scalars().first()
            if cell:
                cell.output_result = form.output_result or ""
            else:
                cell = ProjectTimelineCell(row_id=form.row_id, dept=form.dept, output_result=form.output_result or "")
                db.session.add(cell)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def import_timeline(self, prod_id: int, file_bytes: bytes, replace: bool = True):
        try:
            if not prod_id:
                return Resp.resp_err(msg=ts("msg_err_param"))
            import io
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb[wb.sheetnames[0]]
            # 合并单元格在 openpyxl 中只有左上角有值，其余为 None。
            # 先把合并区的锚点值填充到该区所有单元格，避免跨多行/列的内容丢失。
            merged_val: dict = {}
            for mr in list(ws.merged_cells.ranges):
                anchor = ws.cell(row=mr.min_row, column=mr.min_col).value
                for rr in range(mr.min_row, mr.max_row + 1):
                    for cc in range(mr.min_col, mr.max_col + 1):
                        merged_val[(rr, cc)] = anchor
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            grid = []
            for rr in range(1, max_row + 1):
                row = []
                for cc in range(1, max_col + 1):
                    row.append(merged_val.get((rr, cc), ws.cell(row=rr, column=cc).value))
                grid.append(row)
            if len(grid) < 3:
                return Resp.resp_err(msg=ts("msg_err_param"))

            # 表头：第1行从第4列起为部门名，映射到固定部门列
            header = grid[0]
            dept_cols: dict[int, str] = {}
            for cidx in range(3, len(header)):
                name = str(header[cidx] or "").strip()
                if name in TIMELINE_DEPTS:
                    dept_cols[cidx] = name

            if replace:
                old_rows = db.session.execute(
                    select(ProjectTimelineRow.id).where(ProjectTimelineRow.prod_id == prod_id)
                ).scalars().all()
                if old_rows:
                    db.session.execute(delete(ProjectTimelineCell).where(ProjectTimelineCell.row_id.in_(old_rows)))
                    db.session.execute(delete(ProjectTimelineRow).where(ProjectTimelineRow.id.in_(old_rows)))

            last_year = ""
            last_month = ""
            sort_order = 0
            imported = 0
            for r in grid[2:]:
                def cell(i):
                    return str(r[i]).strip() if i < len(r) and r[i] is not None else ""
                a, b, c = cell(0), cell(1), cell(2)
                dept_values = {dept: cell(ci) for ci, dept in dept_cols.items()}
                has_dept = any(v for v in dept_values.values())
                has_date = bool(b or c)

                if not (a or has_date or has_dept):
                    continue

                sort_order += 1
                if has_date or has_dept:
                    # 日期行：向下填充年/月
                    if a:
                        last_year = a
                    if b:
                        last_month = b
                    row = ProjectTimelineRow(
                        prod_id=prod_id, year=a or last_year, month=b or last_month,
                        day=c, row_type="date", sort_order=sort_order,
                    )
                    row.id = None
                    db.session.add(row)
                    db.session.flush()
                    for dept, val in dept_values.items():
                        if val:
                            db.session.add(ProjectTimelineCell(row_id=row.id, dept=dept, output_result=val))
                    imported += 1
                else:
                    # 仅 A 列有值：年份行 或 阶段里程碑行
                    rtype = "year" if ("年" in a and len(a) <= 8) else "milestone"
                    if rtype == "year":
                        last_year = a
                    row = ProjectTimelineRow(
                        prod_id=prod_id, row_type=rtype, milestone_text=a, sort_order=sort_order,
                    )
                    row.id = None
                    db.session.add(row)
                    imported += 1
            db.session.commit()
            return Resp.resp_ok(data={"imported": imported})
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def export_timeline(self, prod_id: int, output):
        resp = await self.list_timeline(prod_id)
        rows = (resp.data or {}).get("rows") or []

        wb = Workbook()
        ws = wb.active
        ws.title = "时间逻辑线"
        FONT_NAME = "微软雅黑"
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap = Alignment(vertical="top", wrap_text=True)
        bold = Font(name=FONT_NAME, bold=True)
        normal = Font(name=FONT_NAME)
        time_hdr_fill = PatternFill("solid", fgColor="F2F2F2")

        # 各部门配色：(表头底色, 列内容底色)
        dept_styles = {
            "产品部": ("F8CBAD", "FCE4D6"),
            "模型部": ("FFE699", "FFF2CC"),
            "数据部": ("A9D08E", "E2EFDA"),
            "产品开发部-开发": ("C6E0B4", "E2EFDA"),
            "产品开发部-测试": ("C6E0B4", "EBF3E6"),
            "质量部": ("E29EDB", "F4CCE8"),
            "三级通用技术文件": ("F4B7DC", "FCE4F0"),
        }
        default_style = ("D9E1F2", "FFFFFF")

        def dept_hdr_fill(dept):
            return PatternFill("solid", fgColor=dept_styles.get(dept, default_style)[0])

        def dept_body_fill(dept):
            return PatternFill("solid", fgColor=dept_styles.get(dept, default_style)[1])

        total_cols = 3 + len(TIMELINE_DEPTS)
        # 第1行：时间(合并 A1:C1) + 各部门
        ws.cell(row=1, column=1, value="时间")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        for i, dept in enumerate(TIMELINE_DEPTS):
            ws.cell(row=1, column=4 + i, value=dept)
        # 第2行：年/月/日 + 输出结果
        ws.cell(row=2, column=1, value="年")
        ws.cell(row=2, column=2, value="月")
        ws.cell(row=2, column=3, value="日")
        for i in range(len(TIMELINE_DEPTS)):
            ws.cell(row=2, column=4 + i, value="输出结果")
        for rr in (1, 2):
            for cc in range(1, total_cols + 1):
                cobj = ws.cell(row=rr, column=cc)
                cobj.alignment = center
                cobj.font = bold
                if cc <= 3:
                    cobj.fill = time_hdr_fill
                else:
                    cobj.fill = dept_hdr_fill(TIMELINE_DEPTS[cc - 4])

        top = Alignment(vertical="top", wrap_text=True)
        ridx = 3
        seq = []  # (excel_row, row_type, year, month)
        for r in rows:
            obj = r.dict() if hasattr(r, "dict") else r
            row_type = obj.get("row_type") or "date"
            if row_type in ("year", "milestone"):
                ws.cell(row=ridx, column=1, value=obj.get("milestone_text") or "")
                ws.merge_cells(start_row=ridx, start_column=1, end_row=ridx, end_column=total_cols)
                ws.cell(row=ridx, column=1).alignment = center
                ws.cell(row=ridx, column=1).font = bold
                seq.append((ridx, row_type, None, None, {}))
            else:
                year = obj.get("year") or ""
                month = obj.get("month") or ""
                day = obj.get("day") or ""
                ws.cell(row=ridx, column=1, value=year)
                ws.cell(row=ridx, column=2, value=month)
                # 纯数字的「日」按数值写入，避免 Excel「文本存数字」绿色三角警告
                ws.cell(row=ridx, column=3, value=int(day) if str(day).strip().isdigit() else day)
                cells = obj.get("cells") or {}
                for i, dept in enumerate(TIMELINE_DEPTS):
                    c = ws.cell(row=ridx, column=4 + i, value=cells.get(dept) or "")
                    c.fill = dept_body_fill(dept)
                for cc in range(1, total_cols + 1):
                    cobj = ws.cell(row=ridx, column=cc)
                    cobj.alignment = wrap
                    cobj.font = normal
                seq.append((ridx, row_type, year, month, cells))
            ridx += 1

        # 连续相同的「年」(A列) /「月」(B列) 纵向合并，靠上对齐（被非日期行打断则分组重置）
        def merge_runs(col, key_fn):
            i = 0
            while i < len(seq):
                if seq[i][1] != "date":
                    i += 1
                    continue
                key = key_fn(seq[i])
                j = i
                while j + 1 < len(seq) and seq[j + 1][1] == "date" and key_fn(seq[j + 1]) == key:
                    j += 1
                if seq[j][0] > seq[i][0]:
                    ws.merge_cells(start_row=seq[i][0], start_column=col, end_row=seq[j][0], end_column=col)
                    ws.cell(row=seq[i][0], column=col).alignment = top
                i = j + 1

        merge_runs(1, lambda t: t[2])            # 年
        merge_runs(2, lambda t: (t[2], t[3]))    # 月（同年同月）

        # 各部门「输出结果」列：连续相同且非空的内容纵向合并、居中
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def merge_dept(col):
            dept = TIMELINE_DEPTS[col - 4]
            i = 0
            while i < len(seq):
                if seq[i][1] != "date":
                    i += 1
                    continue
                val = (seq[i][4] or {}).get(dept) or ""
                if not val:
                    i += 1
                    continue
                j = i
                while j + 1 < len(seq) and seq[j + 1][1] == "date" and ((seq[j + 1][4] or {}).get(dept) or "") == val:
                    j += 1
                if seq[j][0] > seq[i][0]:
                    ws.merge_cells(start_row=seq[i][0], start_column=col, end_row=seq[j][0], end_column=col)
                ws.cell(row=seq[i][0], column=col).alignment = center_wrap
                i = j + 1

        for ci in range(len(TIMELINE_DEPTS)):
            merge_dept(4 + ci)

        # 全表加细边框（行列线）
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for rr in range(1, ridx):
            for cc in range(1, total_cols + 1):
                ws.cell(row=rr, column=cc).border = border

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 6
        from openpyxl.utils import get_column_letter
        for i in range(len(TIMELINE_DEPTS)):
            ws.column_dimensions[get_column_letter(4 + i)].width = 28
        wb.save(output)
        output.seek(0)
