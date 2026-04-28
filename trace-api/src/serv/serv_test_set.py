import logging
import io
import re
from openpyxl import load_workbook
from sqlalchemy import select, delete, func
from sqlalchemy.sql import desc
from sqlalchemy.exc import IntegrityError
from ..obj.vobj_user import UserObj
from ..model.product import Product, UserProd
from ..obj.vobj_test_set import TestSetObj
from ..model.test_set import TestSet
from ..model.test_case import TestCase
from ..obj.tobj_test_set import TestSetForm
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):
    @staticmethod
    def __to_str(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def __fit_varchar(value: str, size: int = 64):
        txt = (value or "").strip()
        return txt[:size] if len(txt) > size else txt

    @staticmethod
    def __match_col(header: str, aliases: list[str]):
        txt = Server.__to_str(header).lower().replace(" ", "")
        return any(alias in txt for alias in aliases)

    @staticmethod
    def __looks_like_case_code(value: str):
        txt = (value or "").strip()
        if not txt:
            return False
        # 常见用例编号如 TU01-001-0001 / TC-001，过滤标题和说明行
        return bool(re.match(r"^[A-Za-z0-9][A-Za-z0-9._-]{1,63}$", txt)) and "-" in txt


    async def __read_excel(self, file):
        try:
            bys = await file.read()
            wb = load_workbook(io.BytesIO(bys))
            ws = wb[wb.sheetnames[0]]
            test_cases = []
            header_row_idx = None
            col_idx = {
                "code": 0,
                "srs_code": 1,
                "test_type": 2,
                "function": 4,
                "description": 5,
                "precondition": 6,
                "test_step": 7,
                "expect": 8,
                "note": 9,
            }

            max_scan_row = min(ws.max_row, 30)
            for ridx in range(1, max_scan_row + 1):
                row_vals = [self.__to_str(v) for v in ws.iter_rows(min_row=ridx, max_row=ridx, values_only=True).__next__()]
                for cidx, cell_text in enumerate(row_vals):
                    if self.__match_col(cell_text, ["case编号", "用例编号", "测试用例编号", "caseid"]):
                        col_idx["code"] = cidx
                        header_row_idx = ridx
                    if self.__match_col(cell_text, ["需求编号", "srscode", "srs编号"]):
                        col_idx["srs_code"] = cidx
                        header_row_idx = ridx
                    if self.__match_col(cell_text, ["测试类型", "testtype"]):
                        col_idx["test_type"] = cidx
                    if self.__match_col(cell_text, ["功能点", "功能", "function"]):
                        col_idx["function"] = cidx
                    if self.__match_col(cell_text, ["描述", "description"]):
                        col_idx["description"] = cidx
                    if self.__match_col(cell_text, ["前置条件", "precondition"]):
                        col_idx["precondition"] = cidx
                    if self.__match_col(cell_text, ["测试步骤", "步骤", "teststep"]):
                        col_idx["test_step"] = cidx
                    if self.__match_col(cell_text, ["预期结果", "expect"]):
                        col_idx["expect"] = cidx
                    if self.__match_col(cell_text, ["备注", "note"]):
                        col_idx["note"] = cidx
                if header_row_idx is not None:
                    break

            data_start_row = (header_row_idx + 1) if header_row_idx else 2
            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                code_raw = self.__to_str(row[col_idx["code"]] if len(row) > col_idx["code"] else None)
                if not code_raw:
                    continue
                # 丢弃标题/说明行，避免导入脏数据导致长度超限
                if not self.__looks_like_case_code(code_raw):
                    continue

                test_case = TestCase(
                    code=self.__fit_varchar(code_raw, 64),
                    srs_code=self.__fit_varchar(self.__to_str(row[col_idx["srs_code"]] if len(row) > col_idx["srs_code"] else None), 64),
                    test_type=self.__fit_varchar(self.__to_str(row[col_idx["test_type"]] if len(row) > col_idx["test_type"] else None), 64),
                    function=self.__to_str(row[col_idx["function"]] if len(row) > col_idx["function"] else None),
                    description=self.__to_str(row[col_idx["description"]] if len(row) > col_idx["description"] else None),
                    precondition=self.__to_str(row[col_idx["precondition"]] if len(row) > col_idx["precondition"] else None),
                    test_step=self.__to_str(row[col_idx["test_step"]] if len(row) > col_idx["test_step"] else None),
                    expect=self.__to_str(row[col_idx["expect"]] if len(row) > col_idx["expect"] else None),
                    note=self.__to_str(row[col_idx["note"]] if len(row) > col_idx["note"] else None),
                )
                test_cases.append(test_case)
            logger.info("read_test_case_excel: sheet=%s, header_row=%s, data_start=%s, parsed_cases=%s", ws.title, header_row_idx, data_start_row, len(test_cases))
            return None, test_cases
        except IntegrityError:
            logger.exception("")
            return ts("msg_dup_row"), None
        except Exception:
            logger.exception("")
            return ts(msg_err_db), None
    

    async def add_test_set(self, form: TestSetForm, file):
        try:
            sql = select(func.count(TestSet.id)).where(TestSet.product_id == form.product_id, TestSet.stage == form.stage)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            
            row = TestSet(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.flush()
            if file:
                err, test_cases = await self.__read_excel(file)
                if err:
                    return Resp.resp_err(msg=err)
                logger.info("test_cases: %s", len(test_cases))
                for test_case in test_cases:
                    test_case.set_id = row.id
                    logger.info("test_case: %s %s %s %s %s %s %s %s %s", 
                                test_case.code,
                                len(test_case.code), 
                                len(test_case.srs_code), 
                                len(test_case.function), 
                                len(test_case.description), 
                                len(test_case.precondition), 
                                len(test_case.test_step), 
                                len(test_case.expect), 
                                len(test_case.note)
                                )
                    db.session.add(test_case)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_test_set(self, id):
        db.session.execute(delete(TestSet).where(TestSet.id == id))
        db.session.execute(delete(TestCase).where(TestCase.set_id == id))
        db.session.commit()
        return Resp.resp_ok()
   
    async def update_test_set(self, form: TestSetForm, file):
        try:
            sql = select(func.count(TestSet.id)).where(TestSet.product_id == form.product_id, TestSet.stage == form.stage, TestSet.id != form.id)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))

            sql = select(TestSet).where(TestSet.id == form.id)
            row:TestSet = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            
            for key, value in form.dict().items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            if file:
                db.session.execute(delete(TestCase).where(TestCase.set_id == row.id))
                err, test_cases = await self.__read_excel(file)
                if err:
                    return Resp.resp_err(msg=err)
                for test_case in test_cases:
                    test_case.set_id = row.id
                    db.session.add(test_case)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def get_test_set(self, id:str):
        sql = select(TestSet).where(TestSet.id == id)
        row = db.session.execute(sql).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        return Resp.resp_ok(data=TestSetObj(**row.dict()))

    async def list_test_set(self, op_user: UserObj, product_id: int = None, stage: str = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 
    
        sql = select(TestSet, Product).outerjoin(Product, TestSet.product_id == Product.id)
        if product_id:
            sql = sql.where(TestSet.product_id == product_id)
        if stage:
            sql = sql.where(TestSet.stage == stage)
        if not product_id and op_user and op_user.id != 1:
            subquery = select(UserProd.product_id).where(UserProd.user_id == op_user.id).scalar_subquery()
            sql = sql.where(Product.id.in_(subquery))

        sql_count = select(func.count()).select_from(sql)
        total = db.session.execute(sql_count).scalars().first()
        sql = sql.offset(page_size * page_index).limit(page_size).order_by(desc(TestSet.create_time))
        rows: list[TestSet] = db.session.execute(sql).all()

        objs = []
        for row, row_prd in rows:
            obj = TestSetObj(**row.dict())
            if row_prd:
                obj.product_name = row_prd.name
                obj.product_version = row_prd.full_version
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))
             