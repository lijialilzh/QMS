import logging
import os
from typing import List, Tuple
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from sqlalchemy.dialects.postgresql import insert as pg_insert
from openpyxl import load_workbook
from ..obj.vobj_user import UserObj
from ..model.product import Product, UserProd
from ..model.haz import Haz
from ..model.prod_haz import ProdHaz
from ..obj.tobj_prod_haz import ProdHazForm, ProdHazsForm
from ..obj.vobj_prod_haz import ProdHazObj
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from .serv_haz import export_columns
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):

    async def add_prod_hazs(self, form: ProdHazsForm):
        try:
            values = []
            for haz_id in form.haz_ids:
                values.append(dict(prod_id=form.prod_id, haz_id=haz_id))
            db.session.execute(pg_insert(ProdHaz).values(values).on_conflict_do_nothing())
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
    
    async def update_prod_haz(self, form: ProdHazForm):
        row = db.session.execute(select(ProdHaz).where(ProdHaz.id == form.id)).scalars().first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        columns = set(["init_rate", "init_degree", "init_level", "cur_rate", "cur_degree", "cur_level", "rcms", "evidence", "situation", "damage", "deal"])
        for key, value in form.dict().items():
            if value is not None and key in columns:
                setattr(row, key, value)
        db.session.commit()
        return Resp.resp_ok()
   
    async def delete_prod_hazs(self, ids: List[str]):
        if ids:
            db.session.execute(delete(ProdHaz).where(ProdHaz.id.in_(ids)))
            db.session.commit()
        return Resp.resp_ok()

    async def list_prod_haz(self, op_user: UserObj, export = False, prod_id: int = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 

        sql = select(ProdHaz, Product, Haz).outerjoin(Product, ProdHaz.prod_id == Product.id).outerjoin(Haz, ProdHaz.haz_id == Haz.id)
        if prod_id:
            sql = sql.where(ProdHaz.prod_id == prod_id)
        if not prod_id and op_user and op_user.id != 1:
            subquery = select(UserProd.product_id).where(UserProd.user_id == op_user.id).scalar_subquery()
            sql = sql.where(Product.id.in_(subquery))
        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()
        sql = sql.order_by(Haz.code)
        rows: List[Tuple[ProdHaz, Product, Haz]] = db.session.execute(sql).all()
        objs = []
        for row, row_prd, row_haz in rows:
            obj = ProdHazObj(**row.dict())
            if row_haz:
                obj.code = row_haz.code
                obj.source = row_haz.source
                obj.event = row_haz.event
                obj.benefit_flag = row_haz.benefit_flag
                obj.category = row_haz.category

                obj.init_rate = row.init_rate or row_haz.init_rate
                obj.init_degree = row.init_degree or row_haz.init_degree
                obj.init_level = row.init_level or row_haz.init_level

                obj.cur_rate = row.cur_rate or row_haz.cur_rate
                obj.cur_degree = row.cur_degree or row_haz.cur_degree
                obj.cur_level = row.cur_level or row_haz.cur_level

                obj.rcms = row.rcms or row_haz.rcms
                obj.evidence = row.evidence or row_haz.evidence
                obj.situation = row.situation or row_haz.situation
                obj.damage = row.damage or row_haz.damage
                obj.deal = row.deal or row_haz.deal
            if row_prd:
                obj.product_name = row_prd.name
                obj.product_version = row_prd.full_version
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))

    async def export_prod_hazs(self, op_user: UserObj, output, *args, **kwargs):
        resp = await self.list_prod_haz(op_user, *args, export=True, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_haz.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 3):
            obj = row.dict()
            for cidx, key in enumerate(export_columns, 1):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
        wb.save(output)
        output.seek(0)
        