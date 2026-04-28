import logging
import io
import os
from openpyxl import load_workbook
from typing import List
from sqlalchemy import select, delete, func, or_
from sqlalchemy.sql import desc
from ..obj.vobj_user import UserObj
from ..model.product import Product, UserProd
from ..model.prod_dhf import ProdDhf
from ..obj.tobj_prod_dhf import ProdDhfForm
from ..obj.vobj_prod_dhf import ProdDhfObj
from ..utils.sql_ctx import db
from ..utils.i18n import ts
from ..obj import Page, Resp
from . import msg_err_db

logger = logging.getLogger(__name__)


class Server(object):
    @staticmethod
    def __to_str(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def __match_col(header: str, aliases: List[str]):
        txt = Server.__to_str(header).lower().replace(" ", "")
        return any(alias in txt for alias in aliases)

    async def add_prod_dhf(self, form: ProdDhfForm):
        try:
            sql = select(func.count(ProdDhf.id)).where(ProdDhf.prod_id == form.prod_id, ProdDhf.code == form.code)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            row = ProdDhf(**form.dict())
            row.id = None
            db.session.add(row)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))

    async def import_prod_dhfs(self, prod_id: int, file):
        try:
            bys = await file.read()
            wb = load_workbook(io.BytesIO(bys), data_only=True)
            # 按约定读取第一个可见工作表（忽略隐藏模板页）
            visible_sheets = [sheet for sheet in wb.worksheets if getattr(sheet, "sheet_state", "visible") == "visible"]
            ws = visible_sheets[0] if visible_sheets else wb[wb.sheetnames[0]]
            affected = 0
            created = 0
            updated_by_code = 0
            updated_by_name = 0

            # 优先按表头名识别列，避免“固定列位”导致识别错列
            header_row_idx = None
            code_col_idx = None
            name_col_idx = None
            max_scan_row = min(ws.max_row, 10)
            for ridx in range(1, max_scan_row + 1):
                row_vals = [self.__to_str(v) for v in ws.iter_rows(min_row=ridx, max_row=ridx, values_only=True).__next__()]
                for cidx, cell_text in enumerate(row_vals):
                    if code_col_idx is None and self.__match_col(cell_text, ["文件编号", "编号", "file_no", "code"]):
                        code_col_idx = cidx
                    if name_col_idx is None and self.__match_col(cell_text, ["文件名称", "名称", "file_name", "name"]):
                        name_col_idx = cidx
                if code_col_idx is not None:
                    header_row_idx = ridx
                    break

            # 兜底兼容旧模板：序号、编号、名称
            if code_col_idx is None:
                code_col_idx = 1
            if name_col_idx is None:
                name_col_idx = 2
            data_start_row = (header_row_idx + 1) if header_row_idx else 2

            logger.info(
                "import_prod_dhfs: file=%s, sheet=%s, header_row=%s, code_col=%s, name_col=%s, data_start_row=%s",
                getattr(file, "filename", ""),
                ws.title,
                header_row_idx,
                code_col_idx,
                name_col_idx,
                data_start_row,
            )

            sample_codes = []

            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                code = self.__to_str(row[code_col_idx] if len(row) > code_col_idx else None)
                name = self.__to_str(row[name_col_idx] if len(row) > name_col_idx else None)
                if not code:
                    continue
                if len(sample_codes) < 5:
                    sample_codes.append(code)
                existed = db.session.execute(
                    select(ProdDhf).where(ProdDhf.prod_id == prod_id, ProdDhf.code == code)
                ).scalars().first()
                if existed:
                    existed.name = name
                    updated_by_code += 1
                else:
                    # 若编号变化但名称相同，则按名称回写编号，避免残留旧编号（如 XX）
                    existed_by_name = None
                    if name:
                        existed_by_name = db.session.execute(
                            select(ProdDhf).where(ProdDhf.prod_id == prod_id, ProdDhf.name == name)
                        ).scalars().first()
                    if existed_by_name:
                        conflict = db.session.execute(
                            select(func.count(ProdDhf.id)).where(
                                ProdDhf.prod_id == prod_id,
                                ProdDhf.code == code,
                                ProdDhf.id != existed_by_name.id
                            )
                        ).scalar() or 0
                        if conflict == 0:
                            existed_by_name.code = code
                            existed_by_name.name = name
                            updated_by_name += 1
                        else:
                            existed_by_name.name = name
                            updated_by_name += 1
                    else:
                        db.session.add(ProdDhf(prod_id=prod_id, code=code, name=name))
                        created += 1
                affected += 1
            logger.info("import_prod_dhfs sample codes: %s", sample_codes)
            db.session.commit()
            return Resp.resp_ok(data={
                "count": affected,
                "created": created,
                "updated_by_code": updated_by_code,
                "updated_by_name": updated_by_name,
            })
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
    
    async def update_prod_dhf(self, form: ProdDhfForm):
        try:
            sql = select(func.count(ProdDhf.id)).where(ProdDhf.prod_id == form.prod_id, ProdDhf.code == form.code, ProdDhf.id != form.id)
            count = db.session.execute(sql).scalar()
            if count > 0:
                return Resp.resp_err(msg=ts("msg_obj_exist"))
            sql = select(ProdDhf).where(ProdDhf.id == form.id)
            row:ProdDhf = db.session.execute(sql).scalars().first()
            if not row:
                return Resp.resp_err(msg=ts("msg_obj_null"))
            for key, value in form.dict().items():
                if key == "id" or value is None:
                    continue
                setattr(row, key, value)
            db.session.commit()
            return Resp.resp_ok()
        except Exception:
            logger.exception("")
            db.session.rollback()
        return Resp.resp_err(msg=ts(msg_err_db))
   
    async def delete_prod_dhf(self, id: int):
        db.session.execute(delete(ProdDhf).where(ProdDhf.id == id))
        db.session.commit()
        return Resp.resp_ok()

    async def delete_prod_dhfs(self, ids: List[int]):
        if not ids:
            return Resp.resp_ok()
        db.session.execute(delete(ProdDhf).where(ProdDhf.id.in_(ids)))
        db.session.commit()
        return Resp.resp_ok()
    
    async def get_prod_dhf(self, id: int):
        sql = select(ProdDhf, Product).outerjoin(Product, ProdDhf.prod_id == Product.id).where(ProdDhf.id == id)
        row:ProdDhf = db.session.execute(sql).first()
        if not row:
            return Resp.resp_err(msg=ts("msg_obj_null"))
        row, row_prod = row
        obj = ProdDhfObj(**row.dict())
        if row_prod:
            obj.product_name = row_prod.name
            obj.product_version = row_prod.full_version
        return Resp.resp_ok(data=obj) 

    async def list_prod_dhf(self, op_user: UserObj, export=False, prod_id: int = None, page_index: int = 0, page_size: int = 10):
        page_index = page_index if page_index >= 0 else 0
        page_size = page_size if page_size > 0 else 10 

        sql = select(ProdDhf, Product).outerjoin(Product, ProdDhf.prod_id == Product.id)
        if prod_id:
            sql = sql.where(ProdDhf.prod_id == prod_id)

        if not prod_id and op_user and op_user.id != 1:
            subquery = select(UserProd.product_id).where(UserProd.user_id == op_user.id).scalar_subquery()
            # 兼容历史数据：若产品未建立 UserProd 关联，也允许创建人看到其DHF。
            sql = sql.where(or_(Product.id.in_(subquery), Product.create_user_id == op_user.id))

        total = 0
        if not export:
            sql_count = select(func.count()).select_from(sql)
            total = db.session.execute(sql_count).scalars().first()

        sql = sql.order_by(ProdDhf.code)
        if not export:
            sql = sql.offset(page_size * page_index).limit(page_size)
        rows: List[ProdDhf, Product] = db.session.execute(sql).all()
        objs = []
        for row, row_prod in rows:
            obj = ProdDhfObj(**row.dict())
            if row_prod:
                obj.product_name = row_prod.name
                obj.product_version = row_prod.full_version
            objs.append(obj)
        return Resp.resp_ok(data=Page(total=total, page_size=page_size, rows=objs, page_index=page_index))
        
    
    export_columns = ["code", "name"]
    
    async def export_prod_dhfs(self, op_user: UserObj, output, *args, **kwargs):
        resp = await self.list_prod_dhf(op_user, export=True, *args, **kwargs)
        rows = resp.data.rows or []

        temp_path = os.path.join(os.path.dirname(__file__), "temp_prod_dhf.xlsx")
        wb = load_workbook(temp_path)
        ws = wb[wb.sheetnames[0]]
        for ridx, row in enumerate(rows, 2):
            obj = row.dict()
            ws.cell(row=ridx, column=1, value=ridx-1)
            for cidx, key in enumerate(self.export_columns, 2):
                value = obj.get(key)
                ws.cell(row=ridx, column=cidx, value=value)
        wb.save(output)
        output.seek(0)
